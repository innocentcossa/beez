import os
from django.shortcuts import render
from django.http import HttpResponse, JsonResponse
from Beezfees.models import *
import datetime
import numbers
import logging
import operator
from operator import itemgetter
from pickle import FALSE
from pathlib import Path
from django.contrib.auth.models import Permission
from django.shortcuts import render
from django.http import HttpResponse
from django.utils.dateparse import parse_datetime
#from numpy.array_api import sort
#from numpy.ma import count
from collections import OrderedDict
from django.contrib import messages
from django.contrib.auth import *
from django.contrib.auth.decorators import login_required
from django.shortcuts import redirect
import json, sys
from datetime import *
from django.utils.timezone import now
from .models import ltb_revenue_object, ltb_session, ltb_ordinance,lbt_payment_type
from django.shortcuts import render, redirect
from .models import ltb_revenue_structure
from .models import FeeStatement
from django.shortcuts import redirect, get_object_or_404
from django.db import connection
import random
import string
from django.db.models import F, Sum
from django.db import IntegrityError
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from datetime import date, datetime
from datetime import date
from datetime import date, timedelta
from django.db import transaction
from django.contrib import messages
from decimal import Decimal
import openpyxl
from django.http import JsonResponse
from django.shortcuts import render
import json
from .models import ltb_revenue_groups, lbt_currency, lbt_customer, ltb_revenue_object,invoice_list_view,account_balances
from datetime import date
from .models import lbt_company, lbt_invoice_dtl, lbt_customer, lbt_remmitance_hdr
from .models import ltb_session, lbt_company, lbt_registration
from django.shortcuts import  redirect
from django.db import transaction
from .models import lbt_invoice_hdr, ltb_revenue_structure,lbt_exchange_rate
import datetime
import traceback
import logging
#import mysql.connector
from django.contrib import messages
from django.db.models import F, Sum
from datetime import date, timedelta
from django.shortcuts import render, get_object_or_404, redirect
from .models import lbt_invoice_hdr
from django.contrib.auth.decorators import login_required
from .models import account_balances
from django.contrib.auth import authenticate, login
from datetime import datetime, date
from django.shortcuts import get_object_or_404


from django.contrib.auth import authenticate, login
from django.http import JsonResponse

def login_user(request):
    if request.method == 'POST':
        try:
            username = request.POST.get('username')
            password = request.POST.get('password')

            if not username or not password:
                raise ValueError("Username and password are required.")

            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                return JsonResponse({'success': True, 'message': 'Login successful'})
            else:
                return JsonResponse({'success': False, 'message': 'Incorrect username or password'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': 'An error occurred during login.', 'error': str(e)})
        
    return render(request, 'login.html')
# Logout
def logoutuser(request):
    return render(request, 'login.html')
    

@login_required
def home(request):
    user = request.user
    dd = User.objects.filter(username=user).first()

    # Get the active session
    
    active_session = ltb_session.objects.filter(col_active=True).first()
    session=active_session.col_session_id

    # Calculate student enrollment count for the active session
    student_enrolment = lbt_registration.objects.filter(col_registration_status=1, col_session_id=session).count()

    # Get the invoice amount for the active session
    inv_amount = lbt_invoice_hdr.objects.filter(col_session_id=active_session.col_session_id).aggregate(total=Sum('col_inv_total'))['total']

    # Get the projections for the active session
    projections = lbt_remmitance_hdr.objects.filter(col_session_id=active_session.col_session_id).aggregate(total=Sum('col_rem_amount'))['total']

    # Fetch the currency
    currency = lbt_currency.objects.first()

    # Get the debtors balance using raw SQL query
    # with connection.cursor() as cursor:
    #     cursor.execute("SELECT SUM(balance) FROM Account_balances")
    #     result = cursor.fetchone()
    #     debtors = result[0] if result else 0

    context = {
        'student_enrolment': student_enrolment,
        'inv_amount': inv_amount,
        'debtors': debtors,
        'currency': currency,
        'projections': projections,
        'User': dd
    }
    return render(request, 'home.html', context)
from django.db.models.functions import Trim


@login_required
def currency(request):
    user = request.user
    dd = User.objects.filter(username=user).first()
    if request.method == 'POST':
        user = request.user
        dd = User.objects.filter(username=user,is_active=1,id=user.id).first()
        user_id=dd.id
        #print(user_id)
       
        try:
            col_curr_name = request.POST.get('col_curr_name')
            col_curr_code = request.POST.get('col_curr_code')
            col_curr_shortcode = request.POST.get('col_curr_shortcode')
            col_curr_symbol = request.POST.get('col_curr_symbol')
            col_active = bool(request.POST.get('col_active', False))

            existing_currency= lbt_currency.objects.filter(col_curr_code=col_curr_code)
            if not existing_currency:

                currency = lbt_currency(
                    col_curr_name=col_curr_name,
                    col_curr_shortcode=col_curr_shortcode,
                    col_curr_code=col_curr_code,
                    col_curr_symbol=col_curr_symbol,
                    col_active=col_active,
                    
                )
                
                currency.save()
                messages.success(request, 'Currency successfully saved.')
        except Exception as e:
            messages.error(request, f'An error occurred while saving the currency: {str(e)}')    
        
    try:
        currency = lbt_currency.objects.all()
    
    except Exception as e:
          messages.error(request, f'An error occurred while retrieving currencies: {str(e)}')
          currency = []
    context = {
        'currency_list': currency,
        'User':dd}

    return render(request, 'Currency.html', context)

@login_required
def ordinance(request):
    user = request.user
    dd = User.objects.filter(username=user).first()
    if request.method == 'POST':
        user = request.user
        dd = User.objects.filter(username=user,is_active=1,id=user.id).first()
     
        user_id=dd.id
      
        try:
            col_ordinance_name = request.POST.get('col_ordinance_name')
            col_effective_date = request.POST.get('col_effective_date')
            col_active = request.POST.get('col_active', False)
            company = lbt_company.objects.first()            
            col_co_code = company.col_co_code
                
 
            ordinance = ltb_ordinance(
                col_ordinance_name=col_ordinance_name,
                col_effective=col_effective_date,  # Update field name to col_effective
                col_active=col_active,
                col_co_code=col_co_code,
    
                
            )
            ordinance.save()
            messages.success(request, 'Ordinance successfully saved.')
        except Exception as e:
            messages.error(request, f'An error occurred while saving the ordinance: {str(e)}')

    try:
        ordinance_details = ltb_ordinance.objects.all()
    except Exception as e:
        messages.error(request, f'An error occurred while retrieving ordinance details: {str(e)}')
        ordinance_details = []

    context = {
        'ordinance_details': ordinance_details,
        'User':dd
    }

    return render(request, 'ordinance.html', context)

@login_required
def single_customer(request):
    user = request.user
    dd = User.objects.filter(username=user).first()
    try:
        customer = lbt_customer.objects.all()

        context = {
            'customer_list': customer,
            'User':dd
        }
        initial_customer = None

        if request.method == 'POST':
            user = request.user
            dd = User.objects.filter(username=user, is_active=1, id=user.id).first()
            user_id = dd.id
            try:
                cust = request.POST.get('col_cust_no')
                col_firstname = request.POST.get('col_firstname')
                col_lastname = request.POST.get('col_lastname')
                col_middlename = request.POST.get('col_middlename')
                col_sex = request.POST.get('col_sex')
                col_phys_add = request.POST.get('col_phys_add')
                col_mail_add = request.POST.get('col_mail_add')
                col_city = request.POST.get('col_city')
                col_email = request.POST.get('col_email')
                col_mobi_num = request.POST.get('col_mobi_num')
                col_telephone = request.POST.get('col_telephone')
                col_grade = request.POST.get('col_grade')

                last_reg = lbt_customer.objects.order_by('-col_cust_no').first()

                if last_reg:
                    last_reg_no = last_reg.col_cust_no
                    prefix = last_reg_no[:3]  # Extract the prefix
                    last_num = int(last_reg_no[3:])  # Extract the numeric part and convert to integer
                    new_num = last_num + 1  # Increment the number
                    current_new_reg = f"{prefix}{new_num:04d}"  # Combine the prefix and incremented number with leading zeros
                else:
                    current_new_reg = None  # Initial customer number if no customers exist in the database

                existing_customer = lbt_customer.objects.filter(col_cust_no=cust).first()
                if existing_customer:
                    # Update the existing customer record
                    existing_customer.col_firstname = col_firstname
                    existing_customer.col_lastname = col_lastname
                    existing_customer.col_middlename = col_middlename
                    existing_customer.col_sex = col_sex
                    existing_customer.col_phys_add = col_phys_add
                    existing_customer.col_mail_add = col_mail_add
                    existing_customer.col_city = col_city
                    existing_customer.col_email = col_email
                    existing_customer.col_mobi_num = col_mobi_num
                    existing_customer.col_telephone = col_telephone
                    existing_customer.col_grade = col_grade
                    existing_customer.save()
                else:
                    # Create a new customer record
                    save_customer = lbt_customer.objects.create(
                        col_cust_no=current_new_reg,
                        col_firstname=col_firstname,
                        col_lastname=col_lastname,
                        col_middlename=col_middlename,
                        col_sex=col_sex,
                        col_phys_add=col_phys_add,
                        col_mail_add=col_mail_add,
                        col_city=col_city,
                        col_email=col_email,
                        col_mobi_num=col_mobi_num,
                        col_telephone=col_telephone,
                        col_grade=col_grade,
                    )

                company = lbt_company.objects.first()
                if company:
                    col_co_code = company.col_co_code
                else:
                    col_co_code = ''

                active_sessions = ltb_session.objects.filter(col_active=True)
                if active_sessions.exists():
                    session = active_sessions.first()
                    session_code = session.col_session_id

                    existing_registration = lbt_registration.objects.filter(col_cust_no=current_new_reg).first()
                    if existing_registration:
                        # Update the existing registration record
                        existing_registration.col_grade = col_grade
                        existing_registration.col_group_code = existing_customer.col_group_code if existing_customer else None
                        existing_registration.col_session_id = session_code
                        existing_registration.col_co_code = col_co_code
                        existing_registration.save()
                    else:
                        # Create a new registration record
                        reg = lbt_registration.objects.create(
                            col_cust_no=current_new_reg,
                            col_grade=col_grade,
                            col_group_code=existing_customer.col_group_code if existing_customer else None,
                            col_session_id=session_code,
                            col_registration_status=False,
                            col_co_code=col_co_code,
                           
                        )
                        print(reg)
                        col_cust_no_list = lbt_registration.objects.values_list('col_cust_no', flat=True).last()
                        print("col_cust_no_list",col_cust_no_list)
                        currency_codes = lbt_currency.objects.values_list('col_curr_code', flat=True)
                        company = lbt_company.objects.first()            
                        col_co_code = company.col_co_code
                        print(currency_codes)
                        
                        for col_cust_no in col_cust_no_list:
                            for currency_code in currency_codes:
                                account_number = (col_cust_no)   # Append currency code to account number
                                print("account_number",account_number)

                                # Check if an account with the same col_cust_no and col_curr_code already exists
                                existing_account = lbt_accounts.objects.filter(col_cust_no=col_cust_no, col_curr_code=currency_code).exists()
                                print("existing_account",existing_account)
                                if existing_account:
                                    print(f"Skipping existing account for customer {col_cust_no}")
                                    continue
                    
                                # Save the account numbers to lbt_accounts table
                                account = lbt_accounts(
                                    col_cust_no=col_cust_no,
                                    col_co_code=col_co_code,
                                    col_account_no=account_number,
                                    col_curr_code=currency_code
                                )
                                account.save()
                                print("account",account)
                messages.success(request, 'Record saved successfully.')
            except Exception as e:
                messages.error(request, f'An error occurred while saving the record: {str(e)}')

    except Exception as e:
        messages.error(request, f'An error occurred while retrieving customer data: {str(e)}')
        context = {
            'customer_list': []
        }

    return render(request, 'single_customer.html', context)
@login_required
def single_customersssssss(request):
    user = request.user
    dd = User.objects.filter(username=user).first()
    try:
        customer = lbt_customer.objects.all()

        context = {
            'customer_list': customer,
            'User':dd
        }

        if request.method == 'POST':
            try:
                col_cust_no = request.POST.get('col_cust_no')
                col_firstname = request.POST.get('col_firstname')
                col_lastname = request.POST.get('col_lastname')
                col_middlename = request.POST.get('col_middlename')
                col_sex = request.POST.get('col_sex')
                col_phys_add = request.POST.get('col_phys_add')
                col_mail_add = request.POST.get('col_mail_add')
                col_city = request.POST.get('col_city')
                col_email = request.POST.get('col_email')
                col_mobi_num = request.POST.get('col_mobi_num')
                col_telephone = request.POST.get('col_telephone')
                col_grade = request.POST.get('col_grade')
                company = lbt_company.objects.first()            
                col_co_code = company.col_co_code
                
                save_customer = lbt_customer(
                    col_cust_no=col_cust_no,
                    col_firstname=col_firstname,
                    col_lastname=col_lastname,
                    col_middlename=col_middlename,
                    col_sex=col_sex,
                    col_co_code=col_co_code,
                    col_phys_add=col_phys_add,
                    col_mail_add=col_mail_add,
                    col_city=col_city,
                    col_email=col_email,
                    col_mobi_num=col_mobi_num,
                    col_telephone=col_telephone,
                    col_grade=col_grade
                )
                save_customer.save()
                
    
                messages.success(request, 'Record saved successfully.')
            except Exception as e:
                messages.error(request, f'An error occurred while saving the record: {str(e)}')

    except Exception as e:
        messages.error(request, f'An error occurred while retrieving customer data: {str(e)}')
        context = {
            'customer_list': []
        }

    return render(request, 'single_customer.html', context)

@login_required
def download_customer_template(request):
    try:
        template_path = os.path.join(settings.BASE_DIR, 'Beezfees', 'templates', 'customer.xlsx')
        if os.path.exists(template_path):
            with open(template_path, 'rb') as file:
                response = HttpResponse(file.read(), content_type='application/vnd.ms-excel')
                response['Content-Disposition'] = 'attachment; filename="customer.xlsx"'
                return response
        else:
            return HttpResponse('Template file does not exist.')
    except Exception as e:
        return HttpResponse('An error occurred while downloading the template file.')

@login_required
def download_manual(request):
    try:
        template_path = os.path.join(settings.BASE_DIR, 'Beezfees', 'templates', 'Manual.docx')
        if os.path.exists(template_path):
            with open(template_path, 'rb') as file:
                #response = HttpResponse(file.read(), content_type='application/vnd.ms-excel')
                response['Content-Disposition'] = 'attachment; filename="BeezFeez User Manual.docx"'
                return response
        else:
            return HttpResponse('Template file does not exist.')
    except Exception as e:
        return HttpResponse('An error occurred while downloading the template file.')


@login_required
def add_school(request):
    user = request.user
    dd = User.objects.filter(username=user).first()
    if request.method == 'POST':
        user = request.user
        dd = User.objects.filter(username=user).first()
        try:
            data = request.POST
            col_co_code = data.get('col_co_code')

            # Check if a school already exists
            existing_school = lbt_company.objects.first()
            if existing_school:
                existing_school.col_co_reg_no = data.get('col_co_reg_no')
                existing_school.col_co_reg_date = data.get('col_co_reg_date')
                existing_school.col_co_name = data.get('col_co_name')
                existing_school.col_co_shortname = data.get('col_co_shortname')
                existing_school.col_co_phy_address = data.get('col_co_phy_address')
                existing_school.col_co_mailadd = data.get('col_co_mailadd')
                existing_school.col_co_telephone = data.get('col_co_telephone')
                existing_school.col_co_mobile_no = data.get('col_co_mobile_no')
                existing_school.col_co_email_add = data.get('col_co_email_add')
                existing_school.save()
                messages.success(request, 'School details successfully updated.')
            else:
                col_co_reg_no = data.get('col_co_reg_no')
                col_co_reg_date = data.get('col_co_reg_date')
                col_co_name = data.get('col_co_name')
                col_co_shortname = data.get('col_co_shortname')
                col_co_phy_address = data.get('col_co_phy_address')
                col_co_mailadd = data.get('col_co_mailadd')
                col_co_telephone = data.get('col_co_telephone')
                col_co_mobile_no = data.get('col_co_mobile_no')
                col_co_email_add = data.get('col_co_email_add')

                school = lbt_company.objects.create(
                    col_co_code=col_co_code,
                    col_co_reg_no=col_co_reg_no,
                    col_co_reg_date=col_co_reg_date,
                    col_co_name=col_co_name,
                    col_co_shortname=col_co_shortname,
                    col_co_phy_address=col_co_phy_address,
                    col_co_mailadd=col_co_mailadd,
                    col_co_telephone=col_co_telephone,
                    col_co_mobile_no=col_co_mobile_no,
                    col_co_email_add=col_co_email_add
                )
                messages.success(request, 'School successfully added.')
        except Exception as e:
            messages.error(request, f'An error occurred while processing the request: {str(e)}')

    try:
        school_details = lbt_company.objects.all()
    except Exception as e:
        messages.error(request, f'An error occurred while retrieving school details: {str(e)}')
        school_details = []

    context = {
        'Details': school_details,
        'User':dd
    }

    return render(request, 'add_school.html', context)
@login_required
def exchange_rate_list(request):
    user = request.user
    dd = User.objects.filter(username=user).first()
    
    context = {'User':dd
    }

    return render(request, 'exchange_rate_list.html', context)

@login_required
def accounting_period(request):
    
    context = {
    }

    return render(request, 'addAccountPeriod.html', context)

@login_required
def revenue_objects(request):
    user = request.user
    dd = User.objects.filter(username=user).first()
    try:
        if request.method == 'POST':
            user = request.user
            dd = User.objects.filter(username=user,is_active=1,id=user.id).first()
            user_id=dd.id
            try:
                col_revenue_code = request.POST.get('col_revenue_code')
                col_revenue_name = request.POST.get('col_revenue_name')
                col_active = request.POST.get('col_active', False)
                company = lbt_company.objects.first()            
                col_co_code = company.col_co_code

                revenue_object = ltb_revenue_object(
                    col_revenue_code=col_revenue_code,
                    col_co_code=col_co_code,
                    col_revenue_name=col_revenue_name,
                    col_active=col_active,
                   
                )

                revenue_object.save()
                messages.success(request, 'Revenue object saved successfully')
            except Exception as e:
                messages.error(request, f'An error occurred while saving the revenue object: {str(e)}')

        revenue = ltb_revenue_object.objects.all()

        context = {
            'revenue_list': revenue,
            'messages': messages.get_messages(request),  # Include messages in the context
            'User':dd
        }
    except Exception as e:
        messages.error(request, f'An error occurred while retrieving revenue objects: {str(e)}')
        context = {
            'revenue_list': [],
            'messages': messages.get_messages(request),  # Include messages in the context
        }

    return render(request, 'revenue_objects.html', context)

@login_required
def school_details_list(request):
    user = request.user
    dd = User.objects.filter(username=user).first()
    data=request.GET
    school_details = {}
    school_details= lbt_company.objects.all()

    context = { 
               'Details': school_details,
               'User':dd
               }
        
    return render(request, 'school_details_list.html', context)

@login_required
def currency_list(request):
    user = request.user
    dd = User.objects.filter(username=user).first()
    currency= lbt_currency.objects.all()

    context = {
        'currency_list': currency,
        'User':dd
    }

    return render(request, 'currency_list.html', context)

@login_required
def revenue_structure_list(request):
    user = request.user
    dd = User.objects.filter(username=user).first()

    rev_structure_list= ltb_revenue_structure.objects.all()

    context = {
        'revenue_structure_list': rev_structure_list,
        'User':dd
    }

    return render(request, 'revenue_structure_list.html', context)
@login_required
def customer_list(request):
    user = request.user
    dd = User.objects.filter(username=user).first()

    customer= lbt_customer.objects.all()
    print(customer)
    context = {
        'customer_list': customer,
        'User':dd
    }
    return render(request, 'customer_list.html', context)

@login_required
def remmitance_header(request):
    user = request.user
    dd = User.objects.filter(username=user).first()

    remmitance= lbt_remmitance_hdr.objects.all()

    context = {
        ' remmitance_header': remmitance,
        'User':dd
    }

    return render(request, 'remmitance_header.html', context)

@login_required
def revenue_group_list(request):
    user = request.user
    dd = User.objects.filter(username=user).first()

    group_list = ltb_revenue_groups.objects.all()

    context = {
        'revenue_group_list': group_list,  # Use the correct variable name here
        'User':dd
    }

    return render(request, 'revenue_group_list.html', context)

@login_required
def invoice_detail(request):
    user = request.user
    dd = User.objects.filter(username=user).first()

    invoice= lbt_invoice_dtl.objects.all()

    context = {
        ' invoice_detail': invoice,
        'User':dd
    }

    return render(request, 'invoice_detail.html', context)

@login_required
def cost_centre_list(request):
    user = request.user
    dd = User.objects.filter(username=user).first()
    cost_centres = ltb_cost_centre.objects.all()

    context = {
        'cost_centre_list': cost_centres,  # Use a different variable name here
        'User':dd
    }

    return render(request, 'cost_centre_list.html', context)


@login_required
def revenue_objects_list(request):
    user = request.user
    dd = User.objects.filter(username=user).first()
    revenue = ltb_revenue_object.objects.all()

    context = {
        'revenue_list': revenue,
        'User':dd
    }

    return render(request, 'revenue_objects_list.html', context)


@login_required
def session_list(request):
    user = request.user
    dd = User.objects.filter(username=user).first()
    sessions = ltb_session.objects.all()

    context = {
        'session_list': sessions,
        'User':dd
    }

    return render(request, 'session_list.html', context)



@login_required
def accounts_list(request):
    user = request.user
    dd = User.objects.filter(username=user).first()
    accounts = lbt_accounts.objects.all()
    
    names = lbt_customer.objects.filter(col_cust_no__in=accounts.values_list('col_cust_no', flat=True))
    currencies = lbt_currency.objects.filter(col_curr_code__in=accounts.values_list('col_curr_code', flat=True))

    context = {
        'accounts_list': accounts,
        'names': names,
        'currencies':currencies,
        'User':dd
    }

    return render(request, 'accounts_list.html', context)


@login_required
def ordinance_details(request):
    user = request.user
    dd = User.objects.filter(username=user).first()
    ordinance_details = ltb_ordinance.objects.all()

    context = {
        'ordinance_details': ordinance_details,
        'User':dd
    }
    print(ordinance_details)
    return render(request, 'ordinance_detail.html', context)



@login_required
def remmitance_detail(request):
    user = request.user
    dd = User.objects.filter(username=user).first()
    remmitance_detail=lbt_remmitance_dtl.objects.all()

    context = {
        'remmitance_detail': remmitance_detail,
        'User':dd
        }

    
    return render(request, 'remmitance_detail.html', context)



@login_required
def revenue_structure(request):
    user = request.user
    dd = User.objects.filter(username=user).first()
    try:
        if request.method == 'POST':
            user = request.user
            dd = User.objects.filter(username=user,is_active=1,id=user.id).first()
            user_id=dd.id
            try:
                col_ref_id = request.POST.get('col_ref_id')
                col_ordinance_code = request.POST.get('col_ordinance_code')
                col_session_id = request.POST.get('col_session_id')
                col_fee_amount = request.POST.get('col_fee_amount')
                col_group_code = request.POST.get('col_group_code')
                col_curr_code = request.POST.get('col_curr_code')
                col_revenue_name = request.POST.get('col_revenue_name')
                company = lbt_company.objects.first()            
                col_co_code = company.col_co_code
                
                # existing_revenue= ltb_revenue_structure.objects.filter(col_ref_id=col_ref_id)
                # if not existing_revenue:
                revenue_structure = ltb_revenue_structure(
                    col_ref_id=col_ref_id,
                    col_ordinance_code=col_ordinance_code,
                    col_group_code=col_group_code,
                    col_co_code=col_co_code,
                    col_curr_code=col_curr_code,
                    col_fee_amount=col_fee_amount,
                    col_revenue_name=col_revenue_name,
                   
                )

                if col_session_id:
                    revenue_structure.col_session_id = col_session_id

                revenue_structure.save()
                messages.success(request, 'Revenue structure successfully created.')
            except Exception as e:
                messages.error(request, f'An error occurred while creating the revenue structure: {str(e)}')

        ordinances = ltb_ordinance.objects.all()
        revenue_objects = ltb_revenue_object.objects.all()
        sessions = ltb_session.objects.all()
        currency = lbt_currency.objects.all()
        revenue_structure = ltb_revenue_structure.objects.all()
        curr = lbt_currency.objects.filter(col_curr_code__in=revenue_structure.values_list('col_curr_code', flat=True))
        ord = ltb_ordinance.objects.filter(col_ordinance_code__in=revenue_structure.values_list('col_ordinance_code', flat=True))
        ssn = ltb_session.objects.filter(col_session_id__in=revenue_structure.values_list('col_session_id', flat=True))
        grp = ltb_revenue_groups.objects.filter(col_group_code__in=revenue_structure.values_list('col_group_code', flat=True))
        context = {
            'ordinances': ordinances,
            'revenue_objects': revenue_objects,
            'sessions': sessions,
            'currency': currency,
            'curr': curr,
            'ord': ord,
            'grp': grp,
            'ssn': ssn,
            'revenue_structure_list': revenue_structure,
            'User':dd
        }
    except Exception as e:
        messages.error(request, f'An error occurred while retrieving revenue structure details: {str(e)}')
        context = {
            'ordinances': [],
            'revenue_objects': [],
            'sessions': [],
            'currency': [],
            'curr': [],
            'ord': [],
            'grp': [],
            'ssn': [],
            'revenue_structure_list': [],
        }

    return render(request, 'revenue_structure.html', context)


@login_required
def session(request):
    user = request.user
    dd = User.objects.filter(username=user).first()
    try:
        if request.method == 'POST':
            user = request.user
            dd = User.objects.filter(username=user,is_active=1,id=user.id).first()
            user_id=dd.id
            try:
                col_session_code = request.POST.get('col_session_code')
                col_session_name = request.POST.get('col_session_name')
                col_session_year = request.POST.get('col_session_year')
                company = lbt_company.objects.first()            
                col_co_code = company.col_co_code
                col_start_date = datetime.strptime(request.POST.get('col_start_date'), '%Y-%m-%d').date()
                col_end_date = datetime.strptime(request.POST.get('col_end_date'), '%Y-%m-%d').date()
                current_date = date.today()

                col_active = 1 if current_date >= col_start_date and col_end_date > current_date else 0
                existing_session= ltb_session.objects.filter(col_session_code=col_session_code,)
                if not existing_session:
                    session = ltb_session(
                        col_session_code=col_session_code,
                        col_session_name=col_session_name,
                        col_session_year=col_session_year,
                        col_co_code=col_co_code,
                        col_end_date=col_end_date,
                        col_start_date=col_start_date,
                        col_active=col_active,
                    
                    )
                    session.save()
                    messages.success(request, 'Saved session successfully.')
            except Exception as e:
                messages.error(request, f'An error occurred while saving the session: {str(e)}')

        sessions = ltb_session.objects.all()

        context = {
            'session_list': sessions,
            'User':dd
        }
    except Exception as e:
        messages.error(request, f'An error occurred while retrieving sessions: {str(e)}')
        context = {
            'session_list': [],
        }

    return render(request, 'session.html', context)

@require_POST
@login_required
def update_session(request):
    session_id = request.POST.get('col_session_code')
    session_name = request.POST.get('col_session_name')
    session_year = request.POST.get('col_session_year')
    
    try:
        session = ltb_session.objects.get(col_session_code=session_id)
        session.col_session_name = session_name
        session.col_session_year = session_year
        session.save()
        messages.success(request, 'session successfully created.') 
        return JsonResponse({'status': 'success'})
    except ltb_session.DoesNotExist:
        return JsonResponse({'status': 'failed', 'msg': 'Session not found'})
    except Exception as e:
        return JsonResponse({'status': 'failed', 'msg': str(e)})

@login_required
def delete_session(request):
    if request.method == 'POST':
        session_id = request.POST.get('col_session_code')
        try:
            session = ltb_session.objects.get(col_session_code=session_id)
            session.delete()
            return JsonResponse({'status': 'success'})
        except ltb_session.DoesNotExist:
            return JsonResponse({'status': 'failed', 'msg': 'Session not found'})
        except Exception as e:
            return JsonResponse({'status': 'failed', 'msg': str(e)})
@login_required
def payment_methods(request):
    user = request.user
    dd = User.objects.filter(username=user).first()
    try:
        if request.method == 'POST':
            user = request.user
            dd = User.objects.filter(username=user,is_active=1,id=user.id).first()
            user_id=dd.id
            
            try:
                col_pay_type_code = request.POST.get('col_pay_type_code')
                col_pay_type_desc = request.POST.get('col_pay_type_desc')
                col_active = request.POST.get('col_active', False)

                # Check if a record with the same col_pay_type_code already exists

                payment = lbt_payment_type(
                    col_pay_type_code=col_pay_type_code,
                    col_pay_type_desc=col_pay_type_desc,
                    col_active=col_active,
    
                    
                )
                payment.save()
                messages.success(request, 'Payment method successfully saved.')
            except Exception as e:
                messages.error(request, f'An error occurred while saving the payment method: {str(e)}')

        pay = lbt_payment_type.objects.all()
        context = {
            'pay': pay,
            'User':dd
        }
    except Exception as e:
        messages.error(request, f'An error occurred while retrieving payment methods: {str(e)}')
        context = {
            'pay': []
        }

    return render(request, 'payment_methods.html', context)
 
@login_required
def revenue_group(request):
    user = request.user
    dd = User.objects.filter(username=user).first()
    try:
        if request.method == 'POST':
            user = request.user
            dd = User.objects.filter(username=user,is_active=1,id=user.id).first()
            user_id=dd.id
            try:
                col_group_code = request.POST.get('col_group_code')
                col_group_name = request.POST.get('col_group_name')
                company = lbt_company.objects.first()            
                col_co_code = company.col_co_code

                group, created = ltb_revenue_groups.objects.get_or_create(
                    col_group_code=col_group_code,
                    col_co_code=col_co_code,
                 
                    defaults={'col_group_name': col_group_name}
                )

                if not created:
                    group.col_group_name = col_group_name
                    group.save()
                    messages.success(request, 'Revenue group successfully updated.')
                else:
                    messages.success(request, 'Revenue group successfully created.')
            except Exception as e:
                messages.error(request, f'An error occurred while saving the revenue group: {str(e)}')

        group_list = ltb_revenue_groups.objects.all()

        context = {
            'revenue_group_list': group_list,
            'User':dd
        }
    except Exception as e:
        messages.error(request, f'An error occurred while retrieving revenue groups: {str(e)}')
        context = {
            'revenue_group_list': [],
        }

    return render(request, 'revenue_group.html', context)

@login_required
def fetch_ordinance_codes(request):
    ordinance_codes = ltb_ordinance.objects.values_list('col_ordinance_code', flat=True)
    
    #print('Ordinance Codes:', ordinance_codes)
    return JsonResponse(list(ordinance_codes), safe=False)

@login_required
def fetch_group_codes(request):
    group_codes= ltb_revenue_groups.objects.values_list('col_group_code', flat=True)
    #print('Group Codes:', group_codes) 
    return JsonResponse(list(group_codes), safe=False)

@login_required
def fetch_revenue_codes(request):
    revenue_name = ltb_revenue_object.objects.values_list('col_revenue_name', flat=True)
    #print('Revenue name:', revenue_name)
    return JsonResponse(list(revenue_name), safe=False)

@login_required
def fetch_session_codes(request):
    session_codes= ltb_session.objects.values_list('col_session_id', flat=True)
    print('Session Codes:', session_codes) 
    return JsonResponse(list(session_codes), safe=False)

@login_required
def fetch_currency_codes(request):
    currency_codes= ltb_currency.objects.values_list('col_curr_code', flat=True)
    print('Currency Codes:', currency_codes) 
    return JsonResponse(list(currency_codes), safe=False)

@login_required
def fetch_payment_methods(request):
    payment = lbt_payment_type.objects.values_list('col_pay_type_desc', flat=True)
    #print('Payment:', payment) 
    return JsonResponse(list(payment), safe=False)


@login_required
def download_customer_list_file(request):
    try:
        template_path = os.path.join(settings.BASE_DIR, 'Beesfees', 'templates', 'customer_balances.xlsx')
        if os.path.exists(template_path):
            with open(template_path, 'rb') as file:
                response = HttpResponse(file.read(), content_type='application/vnd.ms-excel')
                response['Content-Disposition'] = 'attachment; filename="customer_balances.xlsx"'
                return response
        else:
            return HttpResponse('Template file does not exist.')
    except Exception as e:
        return HttpResponse('An error occurred while downloading the template file.')
        
@login_required
def cost_centres(request):
    try:
        if request.method == 'POST':
            user = request.user
            dd = User.objects.filter(username=user,is_active=1,id=user.id).first()
            user_id=dd.id
            
            try:
                col_centre_code = request.POST.get('col_centre_code')
                col_centre_name = request.POST.get('col_centre_name')
                col_active = request.POST.get('col_active', False)
                company = lbt_company.objects.first()            
                col_co_code = company.col_co_code

                # Create a new ltb_cost_centres object
                cost_centre = ltb_cost_centres(
                    col_centre_code=col_centre_code,           
                    col_co_code = col_co_code,
                    col_centre_name=col_centre_name,
                    col_active=col_active,
                  
                )
                # Save the object to the database
                cost_centre.save()

                messages.success(request, 'Cost centre successfully saved.')
            except Exception as e:
                messages.error(request, f'An error occurred while saving the cost centre: {str(e)}')

        cost_centres = ltb_cost_centres.objects.all()

        context = {
            'cost_centre_list': cost_centres,  # Use a different variable name here
            'User':dd
        }
    except Exception as e:
        messages.error(request, f'An error occurred while retrieving cost centres: {str(e)}')
        context = {
            'cost_centre_list': [],
        }

    return render(request, 'cost_centres.html', context)

@login_required
def chat_of_accounts(request):
    
    context = {
    }

    return render(request, 'chat_of_accounts.html', context)



@login_required
def print_invoice(request, invoice_id):
    # Retrieve the invoice header
    invoice_header = lbt_invoice_hdr.objects.get(col_inv_no=invoice_id)
    
    # Retrieve the invoice details
    invoice_details = lbt_invoice_dtl.objects.filter(col_inv_no=invoice_id)
    
    # Retrieve the company object
    company = lbt_company.objects.first()
    
    # Retrieve the customer object for the remittance header
    customer = lbt_customer.objects.filter(col_cust_no=invoice_header.col_cust_no)
    
    # Prepare a dictionary to store the customer name and last name
    customer_names = {}

    # Iterate over the customer objects and retrieve the first name and last name
    for cust in customer:
        customer_names[cust.col_cust_no] = {
            'first_name': cust.col_firstname,
            'last_name': cust.col_lastname,
            'email':cust.col_email,
            'address':cust.col_phys_add,
            'phone':cust.col_mobi_num
        }
    
    # Combine the invoice header and details into a consolidated invoice object
    consolidated_invoice = {
        'header': invoice_header,
        'details': invoice_details,
        'company': company,
        'customer': customer_names,
    }

    # Pass the consolidated invoice object to the invoice printing template
    return render(request, 'print_invoice.html', {'consolidated_invoice': consolidated_invoice})
@login_required
def fetch_customer_no(request):
    customers = lbt_customer.objects.all()
    data = [
        {
            'col_cust_no': customer.col_cust_no,
            
        }
        for customer in customers
    ]
    return JsonResponse(data, safe=False)

@login_required
def fetch_currency_code(request):
    currency = lbt_currency.objects.values_list('col_curr_code', flat=True)
    print('currency:', currency) 
    return JsonResponse(list(currency), safe=False)



@login_required
def accounts(request):
    user = request.user
    dd = User.objects.filter(username=user).first()

    col_cust_no_list = lbt_customer.objects.values_list('col_cust_no', flat=True)
    currency_codes = lbt_currency.objects.values_list('col_curr_code', flat=True)
    company = lbt_company.objects.first()            
    col_co_code = company.col_co_code
    print(currency_codes)
    
    for col_cust_no in col_cust_no_list:
        for currency_code in currency_codes:
            account_number = (col_cust_no)   # Append currency code to account number

            # Check if an account with the same col_cust_no and col_curr_code already exists
            existing_account = lbt_accounts.objects.filter(col_cust_no=col_cust_no, col_curr_code=currency_code).exists()

            if existing_account:
                print(f"Skipping existing account for customer {col_cust_no}")
                continue
 
            # Save the account numbers to lbt_accounts table
            account = lbt_accounts(
                col_cust_no=col_cust_no,
                col_co_code=col_co_code,
                col_account_no=account_number,
                col_curr_code=currency_code
            )
            account.save()
    messages.success(request, 'Accounts successfully created.')
    
    accounts = lbt_accounts.objects.all()
    
    names = lbt_customer.objects.filter(col_cust_no__in=accounts.values_list('col_cust_no', flat=True))
    currencies = lbt_currency.objects.filter(col_curr_code__in=accounts.values_list('col_curr_code', flat=True))

    context = {
        'accounts_list': accounts,
        'names': names,
        'currencies':currencies,
        'User':dd
    }
    return render(request, 'accounts_list.html', context)


@login_required
def fetch_revenue_groups(request):
    fee_groups = ltb_revenue_groups.objects.all()
    data = [
        {
            'value': fee_group.pk,
            'text': fee_group.col_group_name,
        }
        for fee_group in fee_groups
    ]
    return JsonResponse(data, safe=False)

@login_required
def fetch_revenue_groups2(request):
    products = ltb_revenue_structure.objects.all()
    data = [
        {
            'value': product.pk,
            'text': product.col_group_code,
        }
        for product in products
    ]
    context = {
        'products': products,  # Pass the 'products' variable to the template context
    }
    return JsonResponse(data, safe=False)

@login_required
def fetch_revenue_structure2(request):
    col_revenue_code = request.GET.get('col_revenue_code')
    
    try:
        revenue_structure = lbt_revenue_structure.objects.all()
        response = {
            'status': 'success',
            'revenue_code': revenue_structure.col_revenue_code,
            'fee_amount': revenue_structure.col_fee_amount,
            'currency': revenue_structure.currency,
        }
    except revenue_code.DoesNotExist:
        response = {
            'status': 'failed',
            'msg': 'Revenue structure not found for the specified group code.',
        }
    
    return JsonResponse(response)

@login_required
def fetch_records(request):
    group_code = request.POST.get('group_code')
    records = ltb_revenue_structure.objects.filter(group_code=col_group_code).values()
    print(records)
    return JsonResponse({'records': list(records)})


@login_required
def fetch_records(request):
    if request.method == 'POST':
        group_code = request.POST.get('groupCode')  # Assuming the group code is sent as 'groupCode' parameter in the AJAX request
        try:
            records = ltb_revenue_structure.objects.get(col_group_code=group_code)
            # Retrieve the additional information for the revenue structure object
            description = records.col_revenue_code
            fee_amount = records.col_fee_amount
            qty = records.qty  # Assuming a 'qty' field exists in the model
            total = records.total  # Assuming a 'total' field exists in the model

            data = {
                'Description': description,
                'FeeAmount': fee_amount,
                'Qty': qty,
                'Total': total
            }
            return JsonResponse(data)
        except ltb_revenue_structure.DoesNotExist:
            return JsonResponse({'error': 'Revenue structure not found.'}, status=404)
    else:
        return JsonResponse({'error': 'Invalid request method.'}, status=400)
    
@login_required
def group(request):
    user = request.user
    dd = User.objects.filter(username=user).first()
    try:
        if request.method == 'POST':
            
            try:
                grade_filter = request.POST.get('grade_filter')
                group_filter = request.POST.get('group_filter')

                customers = lbt_registration.objects.filter(col_grade=grade_filter)
                group_codes = ltb_revenue_groups.objects.all()
                grades = lbt_registration.objects.values_list('col_grade', flat=True).distinct()
                revenue_codes = ltb_revenue_structure.objects.filter(col_group_code=group_filter)
                names = lbt_customer.objects.filter(col_cust_no__in=customers.values_list('col_cust_no', flat=True))
                # Move the selected group to the beginning of the group codes list
                group_codes_list = list(group_codes)
                group_codes_list.sort(key=lambda x: x == group_filter, reverse=True)

                context = {
                    'names':names,
                    'customers': customers,
                    'group_codes': group_codes_list,
                    'grades': grades,
                    'selected_grade': grade_filter,
                    'selected_group': group_filter,
                    'User':dd
                }
                return render(request, 'assign_group.html', context)
            except Exception as e:
                messages.error(request, f'An error occurred while processing the request: {str(e)}')

        grades = lbt_registration.objects.values_list('col_grade', flat=True).order_by('col_grade').distinct()
        group_codes = ltb_revenue_groups.objects.all()

        context = {
            'grades': grades,
            'group_codes': group_codes,
            'User':dd
        }
    except Exception as e:
        messages.error(request, f'An error occurred while retrieving data: {str(e)}')
        context = {}

    return render(request, 'assign_group.html', context)


@login_required
def save_group_codes(request):
    user = request.user
    dd = User.objects.filter(username=user).first()
    if request.method == 'POST':
        try:
            selected_grade = request.POST.get('selected_grade')
            customers = lbt_registration.objects.filter(col_grade=selected_grade)
            #assigned_customers=lbt_registration.objects.filter()
            
            for customer in customers:
                group_code_value = request.POST.get(f'group_code{customer.col_cust_no}')
                customer.col_group_code = group_code_value
                customer.save()
            
            # Update registration status
            lbt_registration.objects.filter(col_group_code__isnull=False).update(col_registration_status=True)
            
            context = {
                'customers': customers,
                'User':dd
            }
        
            messages.success(request, 'Fees group assigned successfully.')
            return render(request, 'assign_group.html', context)
        
        except Exception as e:
            messages.error(request, f'An error occurred while retrieving data: {str(e)}')
    
    return redirect('Beezfees:group')


@login_required
def save_invoice(request):
    user = request.user
    dd = User.objects.filter(username=user).first()
    try:
        if request.method == 'POST':
            grade_filter = request.POST.get('grade_filter')
            term_filter = request.POST.get('term_filter')
                             
            customers = lbt_registration.objects.filter(col_grade=grade_filter, col_registration_status=1)
            grades = lbt_registration.objects.values_list('col_grade', flat=True).distinct()
            names = lbt_customer.objects.filter(col_cust_no__in=customers.values_list('col_cust_no', flat=True))
            session = lbt_registration.objects.filter(col_session_id=term_filter)
            year = ltb_session.objects.all()
            term = session.filter(col_session_id=term_filter, col_registration_status=1)
            term = term.filter(col_grade=grade_filter)
            
            # Get session code, session id, and session year from the session table
            try:
                session_data = ltb_session.objects.get(col_session_id=term_filter)
                session_code = session_data.col_session_code
                session_id = session_data.col_session_id
                session_year = session_data.col_session_year

                # Concatenate the session code, session id, and session year
                session_info = f"{session_code} - {session_id} ({session_year})"
            except ltb_session.DoesNotExist:
                session_info = "N/A"

            terms = lbt_registration.objects.values_list('col_session_id', flat=True).distinct()
            context = {
                'customers': term,
                'grades': grades,
                'names': names,
                'session': session,
                'year': year,
                'selected_grade': grade_filter,
                'term': term,
                'terms': terms,
                'selected_term': term_filter,
                'User':dd
            }
            
            return render(request, 'inv.html', context)

    except Exception as e:
        messages.error(request, f"An error occurred: {str(e)}")

    # Handle the case when the request method is not POST or any other exception occurs
    try:
        grades = lbt_registration.objects.values_list('col_grade', flat=True).order_by('col_grade').distinct()
        context = {
            'grades': grades,
            'User':dd
        }
    
        return render(request, 'inv.html', context)
    except Exception as e:
        messages.error(request, f"An error occurred: {str(e)}")
        return render(request, 'inv.html', context)
    
@login_required
def save_invoices(request):
    if request.method == 'POST':
        user = request.user
        dd = User.objects.filter(username=user,is_active=1,id=user.id).first()
     
        user_id=dd.id
        print("USER",user_id)
        
        try:
            selected_customer_count = 0 
            selected_customers = request.POST.getlist('selected_customers[]')
            success_count = 0
            error_count = 0
            # Fetch company details
            company = lbt_company.objects.first()
            print("COMPANY",selected_customers)
            
            col_co_name = company.col_co_name
            col_co_code = company.col_co_code

            # Set invoice dates
            col_inv_date = date.today()
            col_due_date = col_inv_date + timedelta(days=30)

            # Generate invoice number prefix based on account number
            last_invoice = lbt_invoice_hdr.objects.order_by('-col_inv_no').first()
            if last_invoice:
                last_inv_number = int(last_invoice.col_inv_no[3:])
                new_inv_number = last_inv_number + 1
            else:
                new_inv_number = 1

            with transaction.atomic():
                # Iterate over selected customers
                for index, customer_data in enumerate(selected_customers, start=1):
                    col_cust_no, col_group_code,col_session_id, = customer_data.split('|')
                    #print(customer_data)

                    # Fetch accounts for the customer and currency codes
                    accounts = lbt_accounts.objects.filter(col_cust_no=col_cust_no)
                    print("ACCOUNTS",accounts)
                    

                    # Iterate over accounts
                    for account in accounts:
                        currency_code = account.col_curr_code
                        print("CURRENCY CODE",currency_code)
                        col_account_no = account.col_account_no
                        
                        # Determine invoice number prefix based on account number
                        col_inv_no_prefix = 'INV'
                        col_inv_no = f"{col_inv_no_prefix}{new_inv_number:03}"

                        # Check if the generated invoice number already exists
                        while lbt_invoice_hdr.objects.filter(col_inv_no=col_inv_no).exists():
                            new_inv_number += 1
                            col_inv_no = f"{col_inv_no_prefix}{new_inv_number:03}"

                        line_total = 0
                        line_num = 0
                        inv_total = 0

                        # Fetch revenue structures based on the group code and currency code
                        revenue_structures = ltb_revenue_structure.objects.filter(
                            col_curr_code=currency_code, col_group_code=col_group_code
                        )
                        # Calculate line_total and inv_total
                        for revenue_structure in revenue_structures:
                            revenue_name = revenue_structure.col_revenue_name
                            print("REVENUE NAME",revenue_name)
                            fee_amount = revenue_structure.col_fee_amount
                            col_quantity = 1
                            line_total += fee_amount * col_quantity
                            inv_total += fee_amount

                        if line_total == 0:
                            continue
                        
                        
                        existing_invoice=lbt_invoice_hdr.objects.filter(col_cust_no=col_cust_no,col_session_id=col_session_id)
                        if existing_invoice.exists():
                            continue
                        else:
                            
                        # Create invoice header
                            invoice_hdr = lbt_invoice_hdr.objects.create(
                                col_inv_no=col_inv_no,
                                col_co_name=col_co_name,
                                col_co_code=col_co_code,
                                col_cust_no=col_cust_no,
                                col_group_code=col_group_code,
                                col_session_id=col_session_id,
                                col_inv_date=col_inv_date,
                                col_due_date=col_due_date,
                                col_acc_no=col_account_no,  # Associate the account number with the invoice hdr
                                col_curr_code=currency_code,  # Associate the currency code with the invoice hdr
                                col_inv_total=line_total,
                            
                            )
                            print(invoice_hdr)

                            line_num = 0

                            # Create invoice detail
                            for revenue_structure in revenue_structures:
                                revenue_name = revenue_structure.col_revenue_name
                                fee_amount = revenue_structure.col_fee_amount
                                col_quantity = 1
                                line_total = fee_amount * col_quantity
                                line_num += 1

                                if line_total == 0:
                                    continue

                                invoice_dtl = lbt_invoice_dtl.objects.create(
                                    col_inv_no=col_inv_no,
                                    col_co_code=col_co_code,
                                    col_inv_line_no=line_num,
                                    col_revenue_name=revenue_name,
                                    col_fee_amount=fee_amount,
                                    col_line_total=line_total,
                                    col_inv_total=line_total,
                                    col_quantity=1,
                                    
                                )
                               

                            new_inv_number += 1
                            invoice_hdr.col_inv_total = inv_total
                            invoice_hdr.save()
                            success_count += 1

            selected_customer_count = len(selected_customers)
            messages.success(request, f"{success_count} records invoiced successfully.")
            return redirect('Beezfees:invoiced_list')
        except Exception as e:
            error_count += 1
            messages.error(request, f"An error occurred: {str(e)}")

    grades = lbt_customer.objects.values_list('col_grade', flat=True).order_by('col_grade').distinct()
    customers = lbt_invoice_hdr.objects.all()
    names = lbt_customer.objects.filter(col_cust_no__in=customers.values_list('col_cust_no', flat=True))
    currencies = lbt_currency.objects.filter(col_curr_code__in=customers.values_list('col_curr_code', flat=True))
    context = {
        'customers': customers,
        'names': names,
        'currencies': currencies,
        'grades': grades,
        'User':dd
    }
    
    return render(request, 'inv.html', context)

@login_required
def top_navigation(request):
   
    return render(request, 'topNavigation.html')

@login_required
def invoiced_list(request):
    user = request.user
    dd = User.objects.filter(username=user).first()
    invoice = None
    grade_filter = None
    grades = None

    if request.method == 'POST':
        grade_filter = request.POST.get('grade_filter')
        invoice = invoice_list_view.objects.all()
        if grade_filter:
            invoice = invoice.filter(grade=grade_filter)
        grades = lbt_registration.objects.values_list('col_grade', flat=True).distinct()

        print(invoice.count())
        
    context = {
        'invoice': invoice,
        'grade': grade_filter,
        'grades': grades,
        'User':dd
    } 

    return render(request, 'Invoiced_customers_list.html', context)

@login_required
def bulk_invoices(request):
    user = request.user
    dd = User.objects.filter(username=user).first()
    with connection.cursor() as cursor:
        cursor.execute("SELECT DISTINCT Grade FROM invoice_list_view")
        grades = cursor.fetchall()

    if request.method == 'POST':
        grade_filter = request.POST.get('grade_filter')
        with connection.cursor() as cursor:
            if grade_filter:
                cursor.execute("SELECT * FROM invoice_list_view WHERE Grade = %s", [grade_filter])
            else:
                cursor.execute("SELECT * FROM invoice_list_view")
            invoice = cursor.fetchall()
    else:
        with connection.cursor() as cursor:
            cursor.execute("SELECT * FROM invoice_list_view")
            invoice = cursor.fetchall()

    return render(request, 'bulk_invoice.html', {'invoice': invoice, 'grades': grades, 'User':dd})

@login_required
def save_remit(request):
    if request.method == 'POST':
        user = request.user
        dd = User.objects.filter(username=user,is_active=1,id=user.id).first()
        user_id=dd.id
        try:
        # Fetch company details
            company = lbt_company.objects.first()
            col_co_code = company.col_co_code
            
        # Set remittance dates
            col_rem_date = date.today()

        # Retrieve the remittance amount from the request.POST data
            col_rem_amount = request.POST.get('col_rem_amount')
            col_session_id = request.POST.get('col_session_id')
         
        # retrieve currency code
            col_curr_code = request.POST.get('col_curr_code')
            col_acc_no = request.POST.get('col_acc_no')
            col_payment_method = request.POST.get('col_payment_method')
        
            shortcode = request.POST.get('shortcode')
            print(shortcode)
            currency=lbt_currency.objects.all()
            
                
            #if shortcode != col_curr_code:
            #   exchange_rate = lbt_exchange_rate.objects.filter(col_active=True).first()
            #   col_exchange_id = exchange_rate.col_exchange_id
            #  dest_amount = exchange_rate.col_dest_amount
             #    col_rem_amount = Decimal(col_rem_amount)
             #   converted_amount = col_rem_amount/dest_amount

        # Iterate over learner codes and create remittance headers
            learner_codes = request.POST.getlist('col_cust_no')
            invoice_numbers = request.POST.getlist('col_inv_no')
            for cust_no, inv_no in zip(learner_codes, invoice_numbers):
            # Generate remittance number with prefix "REC" and 3 auto-generated digits
                last_remittance = lbt_remmitance_hdr.objects.order_by('-col_rem_no').first()
                if last_remittance:
                    last_rem_number = int(last_remittance.col_rem_no[3:])
                    new_rem_number = last_rem_number + 1
                else:
                     new_rem_number = 1

                col_rem_no = f"REC{new_rem_number:03}"

            # Create remittance header
                rem_hdr = lbt_remmitance_hdr.objects.create(
                    col_rem_no=col_rem_no,
                    col_inv_no=inv_no,
                    col_co_code=col_co_code,
                    col_cust_no=cust_no,
                    col_rem_amount=col_rem_amount,
                    col_session_id=col_session_id,
                    col_rem_date=col_rem_date,
                    col_curr_code=col_curr_code,
                   
                    col_acc_no=col_acc_no,
                    col_payment_method=col_payment_method,
                    
                    )
        except Exception as e:
                messages.error(request, f'An error occurred while saving data: {str(e)}') 
        try:
            company = lbt_company.objects.first()            
            col_co_code = company.col_co_code    
            revenue_structures = ltb_revenue_structure.objects.all()

            revenue_names =''
            fee_amount = 0
            line_num =1  
            line_total = 0
            
            for revenue_structure in revenue_structures:
                fee_amount  = revenue_structure.col_fee_amount
                line_num=1
                
            rem_dtl = lbt_remmitance_dtl.objects.create(
            
                col_inv_no=inv_no,
                col_rem_no=col_rem_no,
                col_rem_amount=col_rem_amount,
                col_rem_lin_no=line_num,
                col_co_code=col_co_code,
                col_cust_no=cust_no, 
                col_rem_line_amount=fee_amount,
                col_rem_date=col_rem_date,
     
                )
            messages.success(request, 'Customer successfully receipted.')
            return redirect('Beezfees:remmittance_list')
        except Exception as e:
            messages.error(request, f'An error occurred while saving data: {str(e)}') 
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT DISTINCT Grade FROM invoice_list_view")
            grades = cursor.fetchall()

        if request.method == 'POST':
            grade_filter = request.POST.get('grade_filter')
            with connection.cursor() as cursor:
                if grade_filter:
                    cursor.execute("SELECT * FROM invoice_list_view WHERE Grade = %s", [grade_filter])
                else:
                    cursor.execute("SELECT * FROM invoice_list_view")
                invoice = cursor.fetchall()
        else:
            with connection.cursor() as cursor:
                cursor.execute("SELECT * FROM invoice_list_view")
                invoice = cursor.fetchall()
    
        return render(request, 'remit_exchange.html', {'invoice': invoice})
    except Exception as e:
        messages.error(request, f"An error occurred: {str(e)}")
        return render(request, 'remit_exchange.html', )
                      
    try:# Fetch invoiced customers for rendering
        customers = lbt_invoice_hdr.objects.all()
        names = lbt_customer.objects.filter(col_cust_no__in=customers.values_list('col_cust_no', flat=True))
        currencies = lbt_currency.objects.filter(col_curr_code__in=customers.values_list('col_curr_code', flat=True))
        context = {
        'invoiced_list': customers,
        'names':  names,
        'currencies': currencies,
        'User':dd
    }
    
        return render(request, 'remit_exchange.html', context)
    except Exception as e:
            messages.error(request, f'An error occurred while retrieving data: {str(e)}')    
    
from django.db import connection
# @login_required
# def account_balances(request):
#     user = request.user
#     dd = User.objects.filter(username=user).first()
#     with connection.cursor() as cursor:
#         cursor.execute("SELECT * FROM Account_balances")
#         balances = cursor.fetchall()
#         cust_nos = [balance[0] for balance in balances]  # Extract the customer numbers from the balances
        
#     names = lbt_customer.objects.filter(col_cust_no__in=cust_nos)
#       #using the 'customers' queryset directly

#     context = {
        
#         'names': names,
#         'balances_list': balances,
#         'User':dd
        
#     }
#     return render(request, 'account_balances.html', context)

@login_required
def balance(request):
    balances = account_balances.objects.all()
    user = request.user
    dd = User.objects.filter(username=user).first()
    
    names = lbt_customer.objects.all()
      #using the 'customers' queryset directly

    context = {
        
        'names': names,
        'balances_list': balances,
        'User':dd
        
    }
    return render(request, 'bal/balance.html', context)


@login_required
def save_remmittances(request):
    user = request.user
    dd = User.objects.filter(username=user,is_active=1,id=user.id).first()
     
    user_id=dd.id
    with connection.cursor() as cursor:
             cursor.execute("SELECT * FROM account_list_view")
             accounts = cursor.fetchall()
    if request.method == 'POST':
        
        
        # Fetch company details
        company = lbt_company.objects.first()
        col_co_code = company.col_co_code
        
        
        # Set remittance dates
        col_rem_date = date.today()

        # Retrieve the remittance amount from the request.POST data
        col_rem_amount = request.POST.get('col_rem_amount')
         
         # retrieve currency code
        col_curr_code= request.POST.get('col_curr_code')
        
        # retrieve customer account
        col_account_no= request.POST.get('col_account_no')
        
        #retrieve customer number
        col_cust_no= request.POST.get('col_cust_no')
        col_payment_method= request.POST.get('col_payment_method')
        
        
            
        # Iterate over learner codes and create remittance headers
        learner_codes = request.POST.getlist('col_cust_no')
        account_numbers = request.POST.getlist('col_account_no')
        for cust_no, inv_no in zip(learner_codes, account_numbers):
            # Generate remittance number with prefix "REC" and 3 auto-generated digits
            last_remittance = lbt_remmitance_hdr.objects.order_by('-col_rem_no').first()
            if last_remittance:
                last_rem_number = int(last_remittance.col_rem_no[3:])
                new_rem_number = last_rem_number + 1
            else:
                new_rem_number = 1

            col_rem_no = f"REC{new_rem_number:03}"
            active_sessions = ltb_session.objects.filter(col_active=True)
            if active_sessions.exists():
                session = active_sessions.first()
                session_code = session.col_session_id

            # Create remittance header
            rem_hdr = lbt_remmitance_hdr.objects.create(
                col_rem_no=col_rem_no,
                col_co_code=col_co_code,
                col_cust_no=col_cust_no,
                
                col_session_id=session_code,
                col_rem_date=col_rem_date,
                col_curr_code=col_curr_code,
                col_acc_no=col_account_no,
                col_payment_method=col_payment_method,
            
                col_rem_amount=col_rem_amount,
                
            )
            
            revenue_structures = ltb_revenue_structure.objects.all()

            
            fee_amount = 0
            line_num =1  
            line_total = 0
            
            for revenue_structure in revenue_structures:
                fee_amount  = revenue_structure.col_fee_amount
                line_num=1
                
            rem_dtl = lbt_remmitance_dtl.objects.create(
            
                col_inv_no=inv_no,
                col_rem_no=col_rem_no,
                col_rem_amount=col_rem_amount,
                col_co_code=col_co_code,
                col_rem_lin_no=line_num,
                col_cust_no=col_cust_no, 
                col_rem_line_amount=fee_amount,
                col_rem_date=col_rem_date,
                
     
                )
            messages.success(request, 'Customer successfully receipted.')
            return redirect('Beezfees:remmittance_list')
         
    
    return render(request, 'accounts_remmittances.html',{'accounts': accounts, 'User':dd})

@login_required
def payment_plan(request):
    user = request.user
    dd = User.objects.filter(username=user).first()
    try:
        if request.method == 'POST':
            user = request.user
            dd = User.objects.filter(username=user,is_active=1,id=user.id).first()
            user_id=dd.id
            try:
                # Retrieve currency code
                col_curr_code = request.POST.get('col_curr_code')

                # Retrieve customer account
                col_account_no = request.POST.get('col_account_no')

                # Set payment plan dates
                col_date_registered = date.today()
                col_expiry_date = request.POST.get('col_expiry_date')
                company = lbt_company.objects.first()            
                col_co_code = company.col_co_code

                # Retrieve customer number
                col_cust_no = request.POST.get('col_cust_no')
                col_plan_amount = request.POST.get('col_plan_amount')
                installments = request.POST.get('col_number_of_installments')

                # Create a new payment plan object
                payment_plan = lbt_payment_plan.objects.create(
                    col_number_of_installments=installments,
                    col_date_registered=col_date_registered,
                    col_cust_no=col_cust_no,        
                    col_co_code =col_co_code,
                    col_plan_amount=col_plan_amount,
                    col_expiry_date=col_expiry_date,
                    col_curr_code=col_curr_code,
                    col_acc_no=col_account_no,
                
                )
                messages.success(request, 'Payment plan successfully registered.')
            except Exception as e:
                messages.error(request, f'An error occurred while registering the payment plan: {str(e)}')

        accounts = lbt_accounts.objects.all()
        names = lbt_customer.objects.filter(col_cust_no__in=accounts.values_list('col_cust_no', flat=True))
        currencies = lbt_currency.objects.filter(col_curr_code__in=accounts.values_list('col_curr_code', flat=True))
        # Prepare the context data for rendering the template
        context = {
            'accounts_list': accounts,
            'names': names,
            'currencies': currencies,
            'User':dd
        }

        # Render the 'payment_plan.html' template with the context data
        return render(request, 'payment_plan.html', context)
    except Exception as e:
        messages.error(request, f'An error occurred while retrieving data: {str(e)}')


@login_required
def download_balance_upload_file(request):
    try:
        account_prefix = request.GET.get('account_prefix')

        # Load the existing Excel sheet
        template_path = os.path.join(settings.BASE_DIR, 'Beezfees', 'templates', 'customer_balances .xlsx')
        if not os.path.exists(template_path):
            return HttpResponse('Template file does not exist.')

        workbook = openpyxl.load_workbook(template_path)
        sheet = workbook.active

        # Retrieve learner information from the database
        customers = lbt_customer.objects.all().distinct().values(
            'col_cust_no', 'col_firstname', 'col_lastname', 'col_middlename'
        )

        customer_account_prefixes = lbt_accounts.objects.all().values_list('col_account_no', flat=True)
        account_prefixes = set([account_no[:3] for account_no in customer_account_prefixes])

        # Filter customers based on the account prefix
        if account_prefix:
            customers = customers.filter(accounts__col_account_no__startswith=account_prefix)

        # Add learner information to the sheet
        for customer in customers:
            customer_data = [
                customer['col_cust_no'],
                customer['col_lastname'],
                customer['col_middlename'],
                customer['col_firstname']
            ]
            cust_no = customer['col_cust_no']
            account_numbers = lbt_accounts.objects.filter(col_cust_no=cust_no).values_list('col_account_no', flat=True)
            for account_number in account_numbers:
                sheet.append(customer_data + [account_number])

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="customer_balances .xlsx"'
        workbook.save(response)
        return response
    except Exception as e:
        return HttpResponse('An error occurred while downloading the template file.')


@login_required
def generate_random_string(length):
    letters = string.digits
    print("LETTERS")
    return ''.join(random.choice(letters) for _ in range(length))

@login_required
def adjustments(request):
    if request.method == 'POST':
        user = request.user
        dd = User.objects.filter(username=user,is_active=1,id=user.id).first()
     
        user_id=dd.id
        # Retrieve the form data
        try:
            col_cust_no = request.POST.get('col_cust_no')
            col_acc_no = request.POST.get('col_acc_no')
            col_inv_no = request.POST.get('col_inv_no')
            col_inv_total = float(request.POST.get('col_inv_total'))  
            adjustment_amount = float(col_inv_total)  
            action = request.POST.get('action')  
            print(action)
        # Determine if the adjustment is a deduction or reversal
            if action=='2':
                col_inv_total = -(col_inv_total)
            else:
                col_inv_total =col_inv_total

        # Retrieve the existing invoice from the database
            existing_invoice = lbt_invoice_hdr.objects.get(col_inv_no=col_inv_no)
            

        # Create a new instance for the adjustment
            adjustment_invoice = lbt_invoice_hdr(
                col_inv_no="ADJ" + random_suffix + "-" + existing_invoice.col_inv_no,
                col_co_name=existing_invoice.col_co_name,
                col_co_code=existing_invoice.col_co_code,
                col_cust_no=col_cust_no,
                col_acc_no=col_acc_no,
                col_session_id=existing_invoice.col_session_id,
                col_group_code=existing_invoice.col_group_code,
                col_inv_date=existing_invoice.col_inv_date,
                col_due_date=existing_invoice.col_due_date,
                col_inv_total=col_inv_total,  # Use the adjusted col_inv_total value
                col_curr_code=existing_invoice.col_curr_code,
               
            )
            adjustment_invoice.save()
       
        # Update the invoice total for the existing invoice
            col_inv_total += adjustment_amount
            existing_invoice.save()
        
        # Create a new instance for the adjustment line item
            adjustment_line_item = lbt_invoice_dtl(
                col_inv_line_no=1,
                col_inv_no=adjustment_invoice.col_inv_no,
                col_revenue_name='ADJUSTMENT',
                col_quantity=1,
                col_fee_amount=adjustment_amount,
                col_line_total=adjustment_amount,
                col_inv_total=col_inv_total,  # Use the adjusted col_inv_total value
                
            )
            adjustment_line_item.save()
        except Exception as e:
           messages.error(request, f'An error occurred while processing the receipt adjustment: {str(e)}') 
        messages.success(request, 'Invoice successfully adjusted.')
        print("existing_invoice.col_inv_total:", existing_invoice.col_inv_total)
        print("adjustment_amount:", adjustment_amount)
            
    #with connection.cursor() as cursor:
     #       cursor.execute("SELECT * FROM invoice_list_view")
      #      invoice = cursor.fetchall()
    # Return a response when the request method is not POST
    #print("before invoice")
    invoice= invoice_list_view.objects.all()
    
    print(invoice.count)
    context= {
    'User':dd
    }
    return render(request, 'adjustments.html',context)


@login_required
def receipting_adjustments(request):
    user = request.user
    dd = User.objects.filter(username=user).first()
    try:
        remmitances = lbt_remmitance_hdr.objects.all()
        #company=lbt_company.objects.first()
        customers = lbt_customer.objects.filter(col_cust_no__in=remmitances.values_list('col_cust_no', flat=True))
        currencies = lbt_currency.objects.filter(col_curr_code__in=remmitances.values_list('col_curr_code', flat=True))
        context = {
            'remmitances': remmitances,
            'customers': customers,
            'currencies': currencies,
            'User':dd
        }

        if request.method == 'POST':
            user = request.user
            dd = User.objects.filter(username=user,is_active=1,id=user.id).first()
            user_id=dd.id
            # Retrieve the form data
            col_cust_no = request.POST.get('col_cust_no')
            col_acc_no = request.POST.get('col_acc_no')
            col_rem_no = request.POST.get('col_rem_no')
            #col_inv_no=request.POST.get('col_inv_no')
            col_rem_total = float(request.POST.get('col_rem_amount'))  # Convert to float
            #col_rem_amount = col_rem_total  # Assign the original col_rem_total value
            col_rem_amount =float(request.POST.get('adjusted_amount') )
            action = request.POST.get('action')  # Assuming the action value is submitted in the form

            # Determine if the adjustment is a deduction or reversal
            if action=='2':
                col_rem_amount = -col_rem_amount # Apply the negative sign to col_rem_amount

            # Retrieve the existing invoice from the database
            existing_receipt = lbt_remmitance_hdr.objects.get(col_rem_no=col_rem_no)
            
            # Generate a unique identifier for the adjusted receipt number
            random_suffix = generate_random_string(3)
            col_rem_date = date.today()
            # Create a new instance for the adjustment
            adjustment = lbt_remmitance_hdr(
                col_rem_no="ADJ" + random_suffix + "-" + existing_receipt.col_rem_no,
                col_inv_no=existing_receipt.col_inv_no,
                col_co_code=existing_receipt.col_co_code,
                col_cust_no=col_cust_no,
                col_acc_no=col_acc_no,
                col_rem_date=col_rem_date,
                col_rem_amount=col_rem_amount ,  # Use the adjusted col_inv_total value
                col_curr_code=existing_receipt.col_curr_code,
               
            )
            adjustment.save()

            # Update the remitance amount for the existing invoice
            #existing_receipt.col_rem_amount += adjustment_amount
            col_rem_date = date.today()
            col_rem_date = date.today()


            #co_rem_line_no=lbt_remmitance_dtl.objects.first()
            # Create a new instance for the adjustment line item
            adjustment_line_item = lbt_remmitance_dtl(
                col_inv_no=existing_receipt.col_inv_no,
                col_rem_lin_no=1,
                col_co_code=existing_receipt.col_co_code,
                col_cust_no=col_cust_no,
                col_rem_no=col_rem_no,
                col_rem_amount=col_rem_amount,
                col_rem_line_amount=col_rem_amount,
                col_rem_date=col_rem_date,
               
            )
            adjustment_line_item.save()
            messages.success(request, 'Receipt successfully adjusted.')

        return render(request, 'receipt_adjustment.html', context)
    
    except Exception as e:
        # Handle the exception
        messages.error(request, f'An error occurred while processing the receipt adjustment: {str(e)}')
        return render(request, 'receipt_adjustment.html')

@login_required
def remmittance_list(request):
    user = request.user
    dd = User.objects.filter(username=user).first()
    remmitances = lbt_remmitance_hdr.objects.all()
    #company=lbt_company.objects.first()
    customers = lbt_customer.objects.filter(col_cust_no__in=remmitances.values_list('col_cust_no', flat=True))
    currencies = lbt_currency.objects.filter(col_curr_code__in=remmitances.values_list('col_curr_code', flat=True))
    context = {
        'remmitances': remmitances,
        'customers': customers,
        'currencies': currencies,
        'User':dd
    }
    
    return render(request, 'remmittance_list.html', context)


@login_required
def generate_random_string(length):
    letters = string.digits
    return ''.join(random.choice(letters) for _ in range(length))

@login_required
def upload_balances(request):
    user = request.user
    dd = User.objects.filter(username=user,is_active=1,id=user.id).first()
     
    user_id=dd.id

    context = {
        'User':dd
    }
    if request.method == 'POST':
        
        file = request.FILES.get('file')
        
        if file:
            # Load the Excel file using openpyxl
            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active

            # Variables to keep track of successful uploads and errors
            success_count = 0
            error_count = 0

            # Iterate over each row and process the data
            for row in sheet.iter_rows(min_row=6, min_col=1, values_only=True):
                # Retrieve the relevant data from the current row
               
                col_cust_no, col_acc_no, col_inv_total = row[0], row[4], row[5]
               

                # Skip saving the record if col_inv_total is 0 or None
                if col_inv_total is None or col_inv_total == 0:
                    continue

                try:
                    company = lbt_company.objects.first()
                    col_co_name = company.col_co_name
                    col_co_code = company.col_co_code
                   
                    active_sessions = ltb_session.objects.filter(col_active=True)
                    if active_sessions.exists():
                        session = active_sessions.first()
                        session_code = session.col_session_id
                        
                    col_inv_date = date.today()
                    col_due_date = col_inv_date + timedelta(days=30)
                  
                # Generate invoice number prefix based on account number
                    last_invoice = lbt_invoice_hdr.objects.order_by('-col_inv_no').first()
                    if last_invoice:
                        last_inv_number = int(last_invoice.col_inv_no[3:])
                        new_inv_number = last_inv_number +1
                        
                    else:
                        new_inv_number = 1
                        
                    col_inv_no_prefix = 'OPB'
                    col_inv_no = f"{col_inv_no_prefix}{new_inv_number:03}"
                    print("invoices", col_inv_no)

                    # Check if the generated invoice number already exists
                    while lbt_invoice_hdr.objects.filter(col_inv_no=col_inv_no).exists():
                        new_inv_number += 1
                        col_inv_no = f"{col_inv_no_prefix}{new_inv_number:03}"
                        print("invoice", col_inv_no)
                                       
                # Create a new instance of lbt_invoice_hdr
                    invoice_hdr = lbt_invoice_hdr(
                        col_inv_no=col_inv_no,
                        col_cust_no=col_cust_no,
                        col_acc_no=col_acc_no,
                        col_inv_total=col_inv_total,
                        col_inv_date=col_inv_date,
                        col_session_id=session_code,
                        col_due_date=col_due_date,
                        col_co_code=col_co_code,
                        col_co_name=col_co_name,
                        
                    )
                    # print("invoice header", invoice_hdr)
                    # Save the instance to the database
                    
                    invoice_hdr.save()
                    print("invoice header", invoice_hdr)
                # print(col_inv_total)    
                    invoice_dtl = lbt_invoice_dtl(
                        col_inv_no=col_inv_no,
                        col_inv_line_no=1,
                        col_co_code=col_co_code,
                        col_fee_amount=col_inv_total,
                        col_inv_total=col_inv_total,
                        
                    )
                    invoice_dtl.save()
                    
                    success_count += 1
                except Exception as e:
                    # Handle any errors that occur during the save operation
                    print("error",e.messages)
                    error_count += 1
                   
            success_message = f"Data uploaded successfully. {success_count} records uploaded."

            if error_count > 0:
                error_message = f"{error_count} records could not be processed due to errors."
                messages.error(request, error_message)
            else:
                messages.success(request, success_message)    
          
    return render(request, 'upload_balances.html', context)
          
@login_required
def sales_report(request):
    invoice = None
    grade_filter = None
    grades = None

    user = request.user
    dd = User.objects.filter(username=user).first()

    if request.method == 'POST':
        grade_filter = request.POST.get('grade_filter')
        invoice = invoice_list_view.objects.all()
        if grade_filter:
            invoice = invoice.filter(grade=grade_filter)
        grades = lbt_registration.objects.values_list('col_grade', flat=True).distinct()

        print(invoice.count())
        
    context = {
        'invoice': invoice,
        'grade': grade_filter,
        'grades': grades,
        'User':dd
    } 

    return render(request, 'sales_report.html', context)


@login_required
def upload_customer(request):
    user = request.user
    dd = User.objects.filter(username=user).first()
    customer = lbt_customer.objects.all()

    context = {
        'customer_list': customer,
        'classes':lbt_level_class.objects.all(),
        'User':dd
    }

    if request.method == 'POST':
        file = request.FILES.get('file')
        grades=request.POST.get('class')
        #return HttpResponse(grade)
        if file:
            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active

            success_count = 0
            error_count = 0

            for row in sheet.iter_rows(min_row=5, min_col=1, values_only=True):
                try:
                    customer_number, firstname, lastname, middlename,gender,grade, physical_address, mail_address, city, email, mobile, telephone, group_code = row
                    grade=request.POST.get('class')
                    # Check if the customer already exists in the database
                    customer, created = lbt_customer.objects.get_or_create(col_cust_no=customer_number)

                    if created:
                        # Set the customer attributes
                        customer.col_firstname = firstname
                        customer.col_lastname = lastname
                        customer.col_middlename = middlename
                        customer.col_sex = gender
                        customer.col_grade = grades
                        customer.col_phys_add = physical_address
                        customer.col_mail_add = mail_address
                        customer.col_city = city
                        customer.col_email = email
                        customer.col_mobi_num = mobile
                        customer.col_telephone = telephone
                        customer.col_group_code = group_code,
                        
                        print(customer)

                        # Save the customer record
                        customer.save()
                        success_count += 1  # Increment the success_count

                        company = lbt_company.objects.first()
                        if company:
                            col_co_code = company.col_co_code
                        else:
                            col_co_code = ''

                        active_sessions = ltb_session.objects.filter(col_active=True)
                        if active_sessions.exists():
                            session = active_sessions.first()
                            session_code = session.col_session_id
                            
                     # Create a registration record
                        reg = lbt_registration.objects.create(
                        
                            col_cust_no=customer_number,
                            col_grade=grade,
                            col_group_code=group_code,
                            col_session_id=session_code,
                            col_registration_status=False,
                            col_co_code=col_co_code
                        )
                        print(reg)

                        col_cust_no_list = lbt_customer.objects.values_list('col_cust_no', flat=True)
                        currency_codes = lbt_currency.objects.values_list('col_curr_code', flat=True)
                        company = lbt_company.objects.first()            
                        col_co_code = company.col_co_code
                        print(currency_codes)
                        
                        for col_cust_no in col_cust_no_list:
                            for currency_code in currency_codes:
                                account_number = (col_cust_no)   # Append currency code to account number

                                # Check if an account with the same col_cust_no and col_curr_code already exists
                                existing_account = lbt_accounts.objects.filter(col_cust_no=col_cust_no, col_curr_code=currency_code).exists()

                                if existing_account:
                                    print(f"Skipping existing account for customer {col_cust_no}")
                                    continue
                    
                                # Save the account numbers to lbt_accounts table
                                account = lbt_accounts(
                                    col_cust_no=col_cust_no,
                                    col_co_code=col_co_code,
                                    col_account_no=account_number,
                                    col_curr_code=currency_code
                                )
                                account.save()

                except Exception as e:
                    # Handle any specific exceptions that may occur during data processing
                    error_count += 1
                    print(f"Error processing row: {row}. Error message: {str(e)}")

            success_message = f"Data uploaded successfully. {success_count} records uploaded."

            if error_count > 0:
                error_message = f"{error_count} records could not be processed due to errors."
                messages.error(request, error_message)
            else:
                messages.success(request, success_message)
    #return render(request, 'learner/import_learner.html', context)
    return render(request, 'upload_customer.html', context)


import decimal

@login_required
def statements(request):
    user = request.user
    dd = User.objects.filter(username=user).first()

   
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT * FROM Statements_View")
            balances = cursor.fetchall()

        return render(request, 'statements.html', {'balances': balances, 'User':dd})
    
    except Exception as e:
        # Handle the exception
        messages.error(request, f'An error occurred while fetching the statements: {str(e)}')
        return render(request, 'statements.html')


@login_required
def fees_statements(request, account_number):
    user = request.user
    dd = User.objects.filter(username=user).first()
    try:
        dat = date.today()
        company = lbt_company.objects.first()

        # Retrieve customer information based on the account number from the view
        with connection.cursor() as cursor:
            cursor.execute("SELECT col_cust_no FROM CRDR_AMOUNTS WHERE col_acc_no = %s LIMIT 1", [account_number])
            col_cust_no = cursor.fetchone()[0]

        # Retrieve customer's first name and last name from lbt_customer model
        try:
            customer = lbt_customer.objects.get(col_cust_no=col_cust_no)
            customer_first_name = customer.col_firstname
            customer_last_name = customer.col_lastname
        except lbt_customer.DoesNotExist:
            customer_first_name = ''
            customer_last_name = ''

        with connection.cursor() as cursor:
            cursor.execute("SELECT * FROM CRDR_AMOUNTS WHERE col_acc_no = %s", [account_number])
            fees_statements = cursor.fetchall()

        running_balance = 0
        processed_statements = []
        

        for statement in fees_statements:
            transaction_date = statement[4]
            description = statement[1]
            
            dr_amount = statement[11]  # Convert to Decimal
            print("dr_amount",dr_amount )
                
            cr_amount = statement[12] # Convert to Decimal
            print("cr_amount",cr_amount )
            

            running_balance += dr_amount - cr_amount

            processed_statement = {
                'transaction_date': transaction_date,
                'description': description,
                'dr_amount': dr_amount,
                'cr_amount': cr_amount,
                'running_balance': running_balance
            }

            processed_statements.append(processed_statement)

        context = {
            'fees_statements': processed_statements,
            'dat': dat,
            'company': company,
            'col_cust_no': col_cust_no,
            'customer_first_name': customer_first_name,
            'customer_last_name': customer_last_name,
            'User':dd
        }
        return render(request, 'fees_statements.html', context)
    
    except Exception as e:
        # Handle the exception
        messages.error(request, f'An error occurred while fetching the fees statements: {str(e)}')
        return render(request, 'fees_statements.html')

@login_required
def receipt(request, cust_num):  
    user = request.user
    dd = User.objects.filter(username=user).first()

    # Retrieve the remittance header
    remittance_header = lbt_remmitance_hdr.objects.filter(col_rem_no=cust_num)

    # Retrieve the customer objects for the remittance header
    customer = lbt_customer.objects.filter(col_cust_no__in=remittance_header.values_list('col_cust_no', flat=True))

    # Prepare a dictionary to store the customer name and last name
    customer_names = {}

    # Iterate over the customer objects and retrieve the first name and last name
    for cust in customer:
        customer_names[cust.col_cust_no] = {
            'first_name': cust.col_firstname,
            'last_name': cust.col_lastname,
            'address':cust.col_phys_add,
            'phone':cust.col_mobi_num
        }
    currency = lbt_currency.objects.filter(col_curr_code__in=remittance_header.values_list('col_curr_code', flat=True))
    currency_names = {}
    for curr in currency:
        currency_names[curr.col_curr_code] = {
            'currency': curr.col_curr_name}
            
    # Retrieve all receipts
    receipts = lbt_remmitance_hdr.objects.all()

    # Retrieve the company object
    company = lbt_company.objects.first()
    
    # Prepare the consolidated receipt object
    consolidated_receipt = {
        'remittance_header': remittance_header,
        'customer': customer_names,
        'company': company,
        'receipts': receipts,
        'currency': curr
    }

    # Pass the consolidated receipt object to the template for rendering
    return render(request, 'receipts.html', {'consolidated_receipt': consolidated_receipt, 'User':dd})


@login_required
def age_analysis(request):
    user = request.user
    dd = User.objects.filter(username=user).first()
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT DISTINCT Grade FROM Account_payment_details")
            grades = [row[0] for row in cursor.fetchall()]  # Fetch unique grades

            grade_filter = request.GET.get('grade_filter')  # Get the grade filter from the request parameters
            days_filter = request.GET.get('days_filter')  # Get the days filter from the request parameters

            query = "SELECT * FROM Account_payment_details WHERE 1=1"
            params = []
            
            if grade_filter:
                query += " AND Grade = %s"
                params.append(grade_filter)

            if days_filter:
                query += " AND days_to_due_date = %s"
                params.append(days_filter)

            cursor.execute(query, params)
            balances = cursor.fetchall()

        context = {
            'balances': balances,
            'grades': grades,
            'selected_grade': grade_filter,
            'days_filter': days_filter,
            'User':dd
        }

    except Exception as e:
        messages.error(request, f'An error occurred while retrieving data: {str(e)}')
    return render(request, 'age_analysis.html', context)

@login_required
def exchange_rates(request):
    user = request.user
    dd = User.objects.filter(username=user).first()
    #currency=  lbt_currency.objects.all()
    if request.method == 'POST':
        user = request.user
        dd = User.objects.filter(username=user,is_active=1,id=user.id).first()
        user_id=dd.id
        col_base_curr = request.POST.get('col_base_curr')
        print(col_base_curr)
        col_base_amount = request.POST.get('col_base_amount')
        col_destination_curr = request.POST.get('col_destination_curr')
        print(col_destination_curr)
        col_exchange_name = request.POST.get('col_exchange_name')
        col_dest_amount = request.POST.get('col_dest_amount')
        col_effective_date  = datetime.strptime(request.POST.get('col_effective_date'), '%Y-%m-%d').date()
        col_expiry_date = datetime.strptime(request.POST.get('col_expiry_date'), '%Y-%m-%d').date()
        company = lbt_company.objects.first()
        col_co_code = company.col_co_code
        current_date = date.today()
    
        if current_date >= col_effective_date and col_expiry_date > current_date:
            col_active = 1
        else:
            col_active = 0

        try:
            exchange_rate = lbt_exchange_rate(
                col_base_curr=col_base_curr,
                col_base_amount=col_base_amount,
                col_exchange_name=col_exchange_name,
                col_destination_curr=col_destination_curr,
                col_dest_amount=col_dest_amount,
                col_co_code=col_co_code,
                col_effective_date=col_effective_date,
                col_expiry_date=col_expiry_date,
                col_active=col_active,
             
            )
            exchange_rate.save()
                
            messages.success(request, 'Exchange rate successfully updated.')
        except Exception as e:
            messages.error(request, 'An error occurred while saving the exchange rate: {}'.format(str(e)))

    exchange_rates = lbt_exchange_rate.objects.all()

    context = {
        "currency":lbt_currency.objects.all(),
        'exchange_rates': exchange_rates,
        'User':dd
    }

    return render(request, 'exchange_rate.html', context)

@login_required
def exchange_rate(request):
    user = request.user
    dd = User.objects.filter(username=user).first()
    exchange_rate = lbt_exchange_rate.objects.all()
    context = {
        'exchange_rate': exchange_rate,
        'User':dd
       }

    return render(request, 'exchange_rate.html', context)

@login_required
def assign_revenue(request):
    user = request.user
    dd = User.objects.filter(username=user).first()
    try:
        context={}
        group = lbt_registration.objects.values_list('col_group_code', flat=True).order_by('col_group_code').distinct()
        ordinances = ltb_ordinance.objects.all()
        revenue_objects = ltb_revenue_object.objects.all()
        sessions = ltb_session.objects.all()
        currency = lbt_currency.objects.all()
        revenue_structure = ltb_revenue_structure.objects.all()
        curr = lbt_currency.objects.filter(col_curr_code__in=revenue_structure.values_list('col_curr_code', flat=True))
        ord = ltb_ordinance.objects.filter(col_ordinance_code__in=revenue_structure.values_list('col_ordinance_code', flat=True))
        ssn = ltb_session.objects.filter(col_session_id__in=revenue_structure.values_list('col_session_id', flat=True))
        grp = ltb_revenue_groups.objects.filter(col_group_code__in=revenue_structure.values_list('col_group_code', flat=True))
        context = {
            'group': group,
            'ordinances': ordinances,
            'revenue_objects': revenue_objects,
            'sessions': sessions,
            'currency': currency,
            'curr': curr,
            'ord': ord,
            'grp': grp,
            'ssn': ssn,
        }

        if request.method == 'POST':
            user = request.user
            dd = User.objects.filter(username=user,is_active=1,id=user.id).first()
            user_id=dd.id
            grade_filter = request.POST.get('grade_filter')
            customers = lbt_registration.objects.filter(col_grade=grade_filter,col_registration_status=1).distinct()
            grades = lbt_registration.objects.values_list('col_grade', flat=True).distinct()
            group_filter = request.POST.get('group_filter')
            session_filter = request.POST.get('session_filter')
            
            terms=lbt_registration.objects.values_list('col_session_id', flat=True).distinct()
            
            
            term=lbt_registration.objects.filter(col_session_id=session_filter)
            revenue_codes = ltb_revenue_structure.objects.filter(col_group_code=group_filter)
            names = lbt_customer.objects.filter(col_cust_no__in= customers.values_list('col_cust_no', flat=True))
            
            try:
                session_data = ltb_session.objects.get(col_session_id=session_filter)
                session_code = session_data.col_session_code
                session_id = session_data.col_session_id
                session_year = session_data.col_session_year

                # Concatenate the session code, session id, and session year
                session_info = f"{session_code} - {session_id} ({session_year})"
            except ltb_session.DoesNotExist:
                session_info = "N/A"

            # Collect the selected customers
            selected_customer=()
            selected_customer_count = 0

            selected_customers = request.POST.getlist('selected_customers[]')

            company = lbt_company.objects.first()
            

            col_co_name = company.col_co_name
            col_co_code = company.col_co_code

            # Set invoice dates
            col_inv_date = date.today()
            col_due_date = col_inv_date + timedelta(days=30)

            # Generate invoice number prefix based on account number
            last_invoice = lbt_invoice_hdr.objects.order_by('-col_inv_no').first()
            if last_invoice:
                last_inv_number = int(last_invoice.col_inv_no[3:])
                new_inv_number = last_inv_number + 1
            else:
                new_inv_number = 1

            for customer_data in selected_customers:
                try:
                    col_cust_no, fee_amount, revenue_name = customer_data.split('|')
                    print('customer_data')
                    col_inv_no_prefix = 'INV'

                    col_inv_no = f"{col_inv_no_prefix}{new_inv_number:03}"

                    line_total = 0
                    line_num = 0
                    inv_total = 0

                    for revenue_structure in customer_data:
                        revenue_name = revenue_structure.col_revenue_name

                        col_quantity = 1
                        line_total += fee_amount * col_quantity
                        inv_total += fee_amount

                    if line_total == 0:
                        continue

                    invoice_hdr = lbt_invoice_hdr.objects.create(
                        col_inv_no=col_inv_no,
                        col_co_name=col_co_name,
                        col_co_code=col_co_code,
                        col_cust_no=col_cust_no,
                        col_group_code=group_code,
                        col_inv_date=col_inv_date,
                        col_due_date=col_due_date,
                        col_acc_no=col_cust_no,  # Associate the account number with the invoice hdr
                        col_curr_code=currency,  # Associate the currency code with the invoice hdr
                        col_inv_total=line_total,
                    
                    )

                    line_num = 0

                    # Create invoice detail
                    for revenue_structure in customer_data:
                        revenue_name = revenue_structure.revenue_name
                        fee_amount = revenue_structure.fee_amount
                        col_quantity = 1
                        line_total = fee_amount * col_quantity
                        line_num += 1
                        print(fee_amount)
                        if line_total == 0:
                            continue

                        invoice_dtl = lbt_invoice_dtl.objects.create(
                            col_inv_no=col_inv_no,
                            col_inv_line_no=line_num,
                            col_revenue_name=revenue_name,
                            col_fee_amount=fee_amount,
                            col_co_code=col_co_code,
                            col_line_total=line_total,
                            col_inv_total=line_total,
                            col_quantity=1,
                          
                        )

                    new_inv_number += 1
                    invoice_hdr.col_inv_total = inv_total
                    invoice_hdr.save()

                except Exception as e:
                    messages.error(request, f'An error occurred while processing the selected customers: {str(e)}')

            context.update({
                'customers': term,
                'revenue_codes': revenue_codes,
                'selected_group': group_filter,
                'revenue_structure_list': revenue_structure.filter(col_group_code=group_filter),
                'grades': grades,
                
                'selected_grade': grade_filter, 
                'names':names,
                'terms':terms,
                'User':dd
            })

    except Exception as e:
        messages.error(request, f'An error occurred: {str(e)}')

    return render(request, 'assign_revenue_objects.html', context)

@login_required
def revenue_invoicing(request):
    try:
        if request.method == 'POST':
            user = request.user
            dd = User.objects.filter(username=user,is_active=1,id=user.id).first()
            user_id=dd.id
            selected_customer_count = 0 
            selected_customers = request.POST.getlist('selected_customers[]')
            
            # Fetch company details
            company = lbt_company.objects.first()
            print(company)
            
            col_co_name = company.col_co_name
            col_co_code = company.col_co_code

            # Set invoice dates
            col_inv_date = date.today()
            col_due_date = col_inv_date + timedelta(days=30)
            # Generate invoice number prefix based on account number
            last_invoice = lbt_invoice_hdr.objects.order_by('-col_inv_no').first()
            if last_invoice:
                last_inv_number = int(last_invoice.col_inv_no[3:])
                new_inv_number = last_inv_number + 1
            else:
                new_inv_number = 1

            # Iterate over selected customers
            for index, customer_data in enumerate(selected_customers, start=1):
                try:
                    col_cust_no, col_group_code,col_session_id, col_revenue_name, col_curr_code, col_fee_amount = customer_data.split('|')

                    # Fetch accounts for the customer and currency codes
                    accounts = lbt_accounts.objects.filter(col_cust_no=col_cust_no)

                    # Iterate over accounts
                    for account in accounts:
                        currency_code = account.col_curr_code
                       
                        col_account_no = account.col_account_no

                        # Determine invoice number prefix based on account number
                        col_inv_no_prefix = 'INV'
                        col_inv_no = f"{col_inv_no_prefix}{new_inv_number:03}"

                        # Check if the generated invoice number already exists
                        while lbt_invoice_hdr.objects.filter(col_inv_no=col_inv_no).exists():
                            new_inv_number += 1
                            col_inv_no = f"{col_inv_no_prefix}{new_inv_number:03}"

                        line_total = 0
                        line_num = 0
                        inv_total = 0

                        # Fetch revenue structures based on the group code and currency code
                        revenue_structures = ltb_revenue_structure.objects.filter(
                            col_curr_code=currency_code, col_group_code=col_group_code
                        )

                        # Calculate line_total and inv_total
                       
                        col_quantity = 1
                        line_total = col_fee_amount * col_quantity
                        inv_total = col_fee_amount

                        if line_total == 0:
                            continue

                        # Create invoice header
                        invoice_hdr = lbt_invoice_hdr.objects.create(
                            col_inv_no=col_inv_no,
                            col_co_name=col_co_name,
                            col_co_code=col_co_code,
                            col_cust_no=col_cust_no,
                            col_group_code=col_group_code,
                            col_session_id=col_session_id,
                            col_inv_date=col_inv_date,
                            col_due_date=col_due_date,
                            col_acc_no=col_account_no,  # Associate the account number with the invoice hdr
                            col_curr_code=currency_code,  # Associate the currency code with the invoice hdr
                            col_inv_total=line_total,
                         
                        )

                        line_num = 0

                        # Create invoice detail
                       
                        col_quantity = 1
                        line_total = col_fee_amount * col_quantity
                        line_num = 1

                        if line_total == 0:
                            continue

                        invoice_dtl = lbt_invoice_dtl.objects.create(
                                col_inv_no=col_inv_no,
                                col_inv_line_no=line_num,
                                col_revenue_name=col_revenue_name,
                                col_fee_amount=col_fee_amount,
                                col_line_total=line_total,
                                col_co_code=col_co_code,
                                col_inv_total=line_total,
                                col_quantity=1,
                                
                                
                            )

                        new_inv_number += 1
                        invoice_hdr.col_inv_total = inv_total
                        invoice_hdr.save()

                except Exception as e:
                    messages.error(request, f'An error occurred while processing the selected customers: {str(e)}')

            selected_customer_count = len(selected_customers)
                    
            messages.success(request, f"{selected_customer_count} records invoiced successfully.")
            return redirect('Beezfees:invoiced_list')

        grades = lbt_customer.objects.values_list('col_grade', flat=True).order_by('col_grade').distinct()
        customers = lbt_invoice_hdr.objects.all()
        names = lbt_customer.objects.filter(col_cust_no__in=customers.values_list('col_cust_no', flat=True))
        currencies = lbt_currency.objects.filter(col_curr_code__in=customers.values_list('col_curr_code', flat=True))
        context = {
            'customers': customers,
            'names': names,
            'currencies': currencies,
            'grades': grades,
            'User':dd
        }
    
    except Exception as e:
         messages.error(request, f'An error occurred: {str(e)}')
                         
    return render(request, 'assign_revenue_objects.html',context)

@login_required
def debtors(request):
    user = request.user
    dd = User.objects.filter(username=user).first()
    try:
          # Extract the customer numbers from the balances
        with connection.cursor() as cursor:
            cursor.execute("SELECT * FROM Account_balances")
            balances = cursor.fetchall()
            print("balances",balances)
            cust_nos = [balance[0] for balance in balances] 
        names = lbt_customer.objects.all()

        context = {
            'names': names,
            'balances_list': account_balances.objects.all(),
            'User':dd
        }

        return render(request, 'debtors.html', context)

    except Exception as e:
    
        messages.error(request, f'An error occurred while fetching the account balances: {str(e)}')
        return render(request, 'debtors.html')


@login_required
def invoice_bulk(request):
    user = request.user
    dd = User.objects.filter(username=user).first()
    # Filter customers by grade
    
    filtered_customers = lbt_customer.objects.filter(col_grade='your_grade_value')
    with connection.cursor() as cursor:
        cursor.execute("SELECT * FROM bulk_invoice_printing")  # Replace your_table_name with the actual table or view name
        invoice_results = cursor.fetchall()

    company = lbt_company.objects.first()

    invoices = []
    for invoice_result in invoice_results:
        col_cust_no = invoice_result[3]  # Assuming col_cust_no is the first field in the result tuple

        # Fetch customer information from lbt_customer
        customer = lbt_customer.objects.get(col_cust_no=col_cust_no)

        # Create a new consolidated_invoice dictionary for each invoice instance
        consolidated_invoice = {
            'company': company,
            'header': invoice_result,
            'customer': {
                'cust_no': col_cust_no,
                'first_name': customer.col_firstname,
                'last_name': customer.col_lastname,
                'email': customer.col_email,
                'address': customer.col_phys_add,
                'phone': customer.col_mobi_num
            },
            'details': []  # Initialize an empty list for the details
        }

        invoices.append(consolidated_invoice)

        # Fetch invoice details from the result tuple
        col_quantity = invoice_result[12]
        col_inv_no=invoice_result[0]
        col_inv_date=invoice_result[6]
        col_due_date=invoice_result[7]
        col_revenue_name = invoice_result[9]
        col_fee_amount = invoice_result[10]
        col_line_total = invoice_result[11]

        # Append the details to the 'details' list in the consolidated_invoice dictionary
        consolidated_invoice['details'].append({
            'col_quantity': col_quantity,
            'col_revenue_name': col_revenue_name,
            'col_fee_amount': col_fee_amount,
            'col_line_total': col_line_total,
            'col_inv_no':col_inv_no,
            'col_inv_date':col_inv_date,
            'col_due_date':col_due_date
            
        })

    return render(request, 'invoice_printing.html', {'invoices': invoices, 'User':dd})


# def register(request):
#     context = {}
#     if request.method=="POST":
#         grade_filter = request.POST.get('grade_filter')
#         term_filter=request.POST.get('term_filter')
#         session_filter = request.POST.get('session_filter')

#         year = ltb_session.objects.all()
#         session = ltb_session.objects.filter(col_session_year= session_filter).first()
#         sessions = ltb_session.objects.values_list('col_session_year', flat=True).distinct()
#         term=lbt_registration.objects.filter(col_session_id=term_filter).first()
#         # terms = lbt_registration.objects.values_list('col_session_id', flat=True).distinct()
#         customers = term.filter(col_grade=grade_filter).all()
#         grades = lbt_registration.objects.values_list('col_grade', flat=True).distinct()
#         names = lbt_customer.objects.filter(col_cust_no__in= customers.values_list('col_cust_no', flat=True))
            
#         try:
#             group= lbt_registration.objects.all()
#             context={
                #context['company'] = company,
#             context['names'] = names
                    # 'year':year
#             context['selected_session'] = session_filter
#             context['session'] = session
#             context['grades'] =grades
#             context['terms'] = terms
#             context['customers'] = customers
#             context['sessions'] = sessions
#             context['selected_grade'] = grade_filter

    #         }
    #     except:
    #         pass

    # return render(request, 'try.html',context)

@login_required
def register(request):
    user = request.user
    dd = User.objects.filter(username=user).first()
    context = {}
    
    if request.method == 'POST':
        try:
            group = lbt_registration.objects.values_list('col_group_code', flat=True).order_by('col_group_code').distinct()

            grade_filter = request.POST.get('grade_filter')
            term_filter=request.POST.get('term_filter')
            session_filter = request.POST.get('session_filter')
            
            session = ltb_session.objects.filter(col_session_year= session_filter)
            sessions = ltb_session.objects.values_list('col_session_year', flat=True).distinct()
            term=lbt_registration.objects.filter(col_session_id=term_filter)
            terms = lbt_registration.objects.values_list('col_session_id', flat=True).distinct()
            customers = term.filter(col_grade=grade_filter)
            year = ltb_session.objects.all()
            grades = lbt_registration.objects.values_list('col_grade', flat=True).distinct()
            names = lbt_customer.objects.filter(col_cust_no__in= customers.values_list('col_cust_no', flat=True))
            try:
                session_data = ltb_session.objects.get(col_session_id=term_filter)
                session_code = session_data.col_session_code
                session_id = session_data.col_session_id
                session_year = session_data.col_session_year

                # Concatenate the session code, session id, and session year
                session_info = f"{session_code} - {session_id} ({session_year})"
            except ltb_session.DoesNotExist:
                session_info = "N/A"

        
        # Collect the selected customers
            selected_customer = ()
            selected_customer_count = 0
            selected_customers = request.POST.getlist('selected_customers[]')

            company = lbt_company.objects.first()
            col_co_name = company.col_co_name
            col_co_code = company.col_co_code

            context['company'] = company
            context['names'] = names
            context['year'] = year
            context['selected_session'] = session_filter
            context['session'] = session
            context['grades'] =grades
            context['terms'] = terms
            context['customers'] = customers
            context['sessions'] = sessions
            context['selected_grade'] = grade_filter
            context['User'] = dd
            
        except Exception as e:
            messagaes.error(request,f'An error occured while retrieving abjects:{str(e)}')
        
    return render(request, 'registration.html', context)

@login_required
def save_register(request):
    if request.method == 'POST':
        user = request.user
        dd = User.objects.filter(username=user,is_active=1,id=user.id).first()
        user_id=dd.id
        try:
            selected_customer_count = 0 
            selected_customers = request.POST.getlist('selected_customers[]')

            for index, customer_data in enumerate(selected_customers, start=1):
                col_cust_no, col_group_code, col_grade, col_session_id, col_co_code = customer_data.split('|')
                
                existing_records = lbt_registration.objects.filter(col_cust_no=col_cust_no, col_session_id=col_session_id)
                
                if existing_records.exists():
                        # Skip saving the record if a duplicate entry is found
                    continue
                
                
                registration = lbt_registration.objects.create(
                col_co_code=col_co_code,
                col_cust_no=col_cust_no,
                col_group_code=col_grade,
                col_grade=col_group_code,
                col_session_id=col_session_id,
                col_registration_status=1,
               
            )
            selected_customer_count = len(selected_customers)
            messages.success(request, f"{selected_customer_count} records registered successfully.")
        except Exception as e:
            messages.error(request, f'An error occurred while saving registration: {str(e)}')
    return render(request, 'registration.html')

@login_required
def registration_list(request):
    user = request.user
    dd = User.objects.filter(username=user).first()
    registered_customers= lbt_registration.objects.all()
     
    names = lbt_customer.objects.filter(col_cust_no__in= registered_customers.values_list('col_cust_no', flat=True))
    context = {
        'customer_list': registered_customers,
        'names':names,
        'User':dd
    }
    return render(request, 'registration_list.html', context)


@login_required
def save_registration(request):
    
    if request.method == 'POST':
        user = request.user
        dd = User.objects.filter(username=user,is_active=1,id=user.id).first()
        user_id=dd.id
        selected_customer_count = 0 
        selected_customers = request.POST.getlist('selected_customers[]')
        col_reg_id = request.POST.get('col_reg_id')
        col_cust_no = request.POST.get('col_cust_no')
        col_grade = request.POST.get('col_grade')
        col_session_id = request.POST.get('col_session_code')
        col_registration_status = request.POST.get('status')
        #if col_registration_status == 'Registered':
        if col_registration_status == 'Register':
            col_registration_status = 1
        else:
            col_registration_status = 0
       # existing_customer = lbt_registration.objects.filter(col_cust_no=col_cust_no, col_session_id=session_code).first()

        
        # Check if col_reg_id exists
        try:
            registration = lbt_registration.objects.get(col_reg_id=col_reg_id)
            # Update the existing record
           
            registration.col_session_id = col_session_id
            registration.col_cust_no = col_cust_no
            registration.col_grade =  col_grade
            registration.col_registration_status = col_registration_status
            
            registration.save()
        except lbt_registration.DoesNotExist:
            # col_reg_id does not exist, create a new record
            registration = lbt_registration.objects.create(
                
                col_session_id=col_session_id,
                col_cust_no=col_cust_no,
                col_grade= col_grade,
                col_registration_status=col_registration_status,
                col_reg_id=col_reg_id,
                
            )
        selected_customer_count =+1
        messages.success(request, f"{selected_customer_count} records changed status successfully.")
    registered_customers= lbt_registration.objects.all()
    names = lbt_customer.objects.filter(col_cust_no__in= registered_customers.values_list('col_cust_no', flat=True))
    
    context = {
        'customer_list': registered_customers,
        'names':names,
        'User':dd
    }
   
    return render(request, 'registration_list.html',context)


@login_required
def save_levels(request):
    user = request.user
    dd = User.objects.filter(username=user).first()
    try:
        if request.method == 'POST':
            
            try:
                
                col_grade = request.POST.get('col_grade')

                grade, created = lbt_progresssion.objects.get_or_create(
                    col_grade=col_grade,

                )

                if not created:
                    grade.col_grade = col_grade
                    grade.save()
                    messages.success(request, ' grade successfully updated.')
                else:
                    messages.success(request, 'grade successfully created.')
            except Exception as e:
                messages.error(request, f'An error occurred while saving the revenue group: {str(e)}')

        grade_list = lbt_progresssion.objects.all()

        context = {
            'grades': grade_list,
            'User':dd
        }
    except Exception as e:
        messages.error(request, f'An error occurred while retrieving revenue groups: {str(e)}')
        

    return render(request, 'save_classes.html',context)

@login_required
def save_grade(request):
    user = request.user
    dd = User.objects.filter(username=user).first()
    try:
        if request.method == 'POST':
            
            try:
                
                col_grade = request.POST.get('col_grade')

                grade, created = lbt_grades.objects.get_or_create(
                    col_grade_name=col_grade,

                )

                if not created:
                    grade.col_grade = col_grade
                    grade.save()
                    messages.success(request, ' grade successfully updated.')
                else:
                    messages.success(request, 'grade successfully created.')
            except Exception as e:
                messages.error(request, f'An error occurred while saving the revenue group: {str(e)}')

        grade_list = lbt_grades.objects.all()

        context = {
            'grades': grade_list,
            'User':dd
        }
    except Exception as e:
        messages.error(request, f'An error occurred while retrieving revenue groups: {str(e)}')
        

    return render(request, 'grades.html',context)

@login_required
def save_classes(request):
    user = request.user
    dd = User.objects.filter(username=user).first()
    #try:
    if request.method=="POST":
        #try:
        col_class_name=request.POST.get("col_class_name")
        col_level=request.POST.get("col_level") 
        col_grade=request.POST.get("col_grade") 


       
        created= lbt_level_class.objects.filter(col_class_name=col_class_name,col_level=col_level,col_grade=col_grade).first()

        if not created:

            level = lbt_level_class(col_class_name=col_class_name,
            col_grade=col_grade,
            col_level=col_level)
            level.save()
            messages.success(request, ' class successfully updated.')
        else:
            messages.success(request, 'class successfully created.')
        
            # except Exception as e:
            #     messages.error(request, f'An error occurred while saving Class: {str(e)}')

   
           
    # except Exception as e:
    #     messages.error(request, f'An error occurred while retrieving a Class: {str(e)}')
    
    class_list=lbt_level_class.objects.all()
            
    context={
        "classes":lbt_level_class.objects.all(),
        "levels":lbt_progresssion.objects.all(),
        "grades":lbt_grades.objects.all() ,
        'User':dd}
    
    return render(request,"save_class.html", context)

@login_required
def filter_records(request):
    context={}
    try:
        
        if request.method == 'POST':
            v_year = request.POST.get("session")
            v_term = request.POST.get("term")
            v_class = request.POST.get('grade')
            v_year="2024"
            v_term ="T01" 
            v_class="3YELLOW"
            print(v_year)
            print(v_term)
            print(v_class)
         
            source_rec = lbt_registration.objects.all()
            filter_rec=source_rec
            print(source_rec.count())
            if v_year != 'All':
                filter_rec=source_rec.objects.filter(col_session_year=v_year)
                print(filter_rec.count())
            else:
                filter_rec=filter_rec
            
            if v_term != 'All':
                filter_rec=filter_rec.filter(col_session_code=v_term)
            else:
                filter_rec=filter_rec
            
            if v_class != 'All':
                filter_rec=filter_rec.filter(col_grade=v_class)
            else:
                filter_rec=filter_rec 
            context={
            'filter_rec':filter_rec
            } 
            customers = filter_rec.count()
            print(customers)
            messages.success(request, f"{ customers } records .")
    except:
        context={}
        
    return render(request, 'progress_classes.html',context) 
       
@login_required           
def progress_class(request):
    user = request.user
    dd = User.objects.filter(username=user).first()
    context = {}
    if request.method =="POST":
        try:

            session_filter = request.POST.get('session_filter')
            grade_filter = request.POST.get('grade_filter')
                        
            grades = lbt_registration.objects.values_list('col_grade', flat=True).distinct()
            
            session = lbt_registration.objects.filter(col_session_id=session_filter)
           
            sessions = lbt_registration.objects.values_list('col_session_id', flat=True).distinct()
            year = ltb_session.objects.all()
           
            customers = session.filter(col_grade=grade_filter) 
            names = lbt_customer.objects.filter(col_cust_no__in= customers.values_list('col_cust_no', flat=True))
            #session_code = ltb_session.objects.filter(col_session_code__in= session.values_list('col_session_code', flat=True))

            try:
                session_data = ltb_session.objects.get(col_session_id=session_filter)
                session_code = session_data.col_session_code
                session_id = session_data.col_session_id
                session_year = session_data.col_session_year

                # Concatenate the session code, session id, and session year
                session_info = f"{session_code} - {session_id} ({session_year})"
            except ltb_session.DoesNotExist:
                session_info = "N/A"
 
            v_year = request.POST.get('session_filter')
            v_class = request.POST.get('grade_filter')
       
            print(v_year)
            #print(v_term)
            print(v_class)
        
            source_rec = lbt_registration.objects.all()
            filter_rec=source_rec
            print(source_rec.count())
            if v_year != 'All':
                filter_rec=lbt_registration.objects.filter(col_session_id=v_year)
                print(filter_rec.count())
            else:
                filter_rec=filter_rec
            
        
            if v_class != 'All':
                filter_rec=filter_rec.filter(col_grade=v_class)
                print(filter_rec.count())
            else:
                filter_rec=filter_rec 
 
            company = lbt_company.objects.first()
            col_co_name = company.col_co_name
            col_co_code = company.col_co_code

            context['company'] = company
            context['names'] = names
            context['year'] = year
            context['selected_grade'] = grade_filter
            context['terms'] = sessions
            context['customers'] = session
            context['grades'] = grades 
            context['User'] = dd
        except Exception as e:
            messages.error(request, f'An error occurred while retrieving sessions: {str(e)}')
        
    return render(request, 'progress_classes.html',context)


@login_required
def save_progression(request):
    if request.method == 'POST':
        user = request.user
        dd = User.objects.filter(username=user,is_active=1,id=user.id).first()
        user_id=dd.id

        context = {'User':dd}
        try:
            selected_customers = request.POST.getlist('selected_customers[]')
            grade = request.POST.get('col_grade')
            group_code = request.POST.get('col_group_code')
            session_year = request.POST.get('col_session_year')
            session_code = request.POST.get('col_session_code')

            session = get_object_or_404(ltb_session, col_session_year=session_year, col_session_code=session_code)
            sessions=session.col_session_id

            for index, customer_data in enumerate(selected_customers, start=1):
                col_cust_no, col_co_code = customer_data.split('|')

                existing_records = lbt_registration.objects.filter(col_cust_no=col_cust_no, col_session_id=sessions)

                if not existing_records:
                    progression = lbt_registration.objects(
                        col_co_code=col_co_code,
                        col_cust_no=col_cust_no,
                        col_group_code=group_code,
                        col_grade=grade,
                        col_session_id=sessions,
                        col_registration_status=1,
                    )
                    progression.save()
                    selected_customer_count = len(selected_customers)
                    messages.success(request, f"{selected_customer_count} records saved successfully.")
                    # Skip saving the record if a duplicate entry is found
                    
                # else:
                #     existing_records. col_cust_no=col_cust_no,
                #     existing_records.col_group_code=group_code,
                #     existing_records.col_grade=grade,
                #     existing_records.col_session_id=sessions,

                #     selected_customer_count = len(selected_customers)
                #     messages.success(request, f"{selected_customer_count} records already moved.")
        except Exception as e:
            messages.error(request, f'An error occurred while saving progression: {str(e)}')

    return render(request, 'progress_classes.html', context)

@login_required
def fetch_grades(request):
    grades = lbt_level_class.objects.values_list('col_class_name', flat=True).distinct()

    return JsonResponse(list(grades), safe=False)

@login_required    
def fetch_group_codes(request):
    group_codes= ltb_revenue_groups.objects.values_list('col_group_code', flat=True)
    return JsonResponse(list(group_codes), safe=False)

@login_required
def fetch_session_year(request):
    academic_year= ltb_session.objects.values_list('col_session_year', flat=True).distinct()
    return JsonResponse(list(academic_year), safe=False)

@login_required
def fetch_session_code(request):
    session_codes= ltb_session.objects.values_list('col_session_code', flat=True).distinct()
    return JsonResponse(list(session_codes), safe=False)

@login_required
def fetch_session_id(request):
    session_id= ltb_session.objects.values_list('col_session_id', flat=True).distinct()
    return JsonResponse(list(session_id), safe=False)







@login_required
def fetch_cur(request):
    cur = lbt_currency.objects.values_list('col_curr_name', flat=True)
    
    return JsonResponse(list(cur), safe=False)

@login_required
def save_remittances(request):
    if request.method == 'POST':
        user = request.user
        dd = User.objects.filter(username=user,is_active=1,id=user.id).first()
        user_id=dd.id
        # print("user_id",user_id)
        try:
        # Fetch company details
            company = lbt_company.objects.first()
            col_co_code = company.col_co_code
            
        # Set remittance dates
            col_rem_date = date.today()
            
            inv_no=None
        
        # Retrieve the remittance amount from the request.POST data
            col_rem_amount = request.POST.get('col_rem_amount')
            #print("col_rem_amount",col_rem_amount)
            col_session_id = request.POST.get('col_session_id')
            print("col_session_id",col_session_id)
         
         # retrieve currency code
            col_curr_code = request.POST.get('col_curr_code')
            print("col_curr_code",col_curr_code)
            col_acc_no = request.POST.get('col_acc_no')
            print("col_acc_no",col_acc_no)
            col_payment_method = request.POST.get('col_payment_method')
           
            
            
        # Iterate over learner codes and create remittance headers
            learner_codes = request.POST.getlist('col_cust_no')
            invoice_numbers = request.POST.getlist('col_inv_no')
            for cust_no, inv_no in zip(learner_codes, invoice_numbers):
            # Generate remittance number with prefix "REC" and 3 auto-generated digits
                last_remittance = lbt_remmitance_hdr.objects.order_by('-col_rem_no').first()
                if last_remittance:
                    last_rem_number = int(last_remittance.col_rem_no[3:])
                    new_rem_number = last_rem_number + 1
                else:
                     new_rem_number = 1

                col_rem_no = f"REC{new_rem_number:03}"

            # Create remittance header
                rem_hdr = lbt_remmitance_hdr.objects.create(
                    col_rem_no=col_rem_no,
                    col_inv_no=inv_no,
                    col_co_code=col_co_code,
                    col_cust_no=cust_no,
                    col_rem_amount=col_rem_amount,
                    col_session_id=col_session_id,
                    col_rem_date=col_rem_date,
                    col_curr_code=col_curr_code,
                    col_acc_no=col_acc_no,
                    col_payment_method=col_payment_method,
                   
                    )
        except Exception as e:
                messages.error(request, f'An error occurred while saving data: {str(e)}') 
        try:
            company = lbt_company.objects.first()            
            col_co_code = company.col_co_code    
            revenue_structures = ltb_revenue_structure.objects.all()

            revenue_names =''
            fee_amount = 0
            line_num =1  
            line_total = 0
            
            for revenue_structure in revenue_structures:
                fee_amount  = revenue_structure.col_fee_amount
                line_num=1
                
            rem_dtl = lbt_remmitance_dtl.objects.create(
            
                col_inv_no=inv_no,
                col_rem_no=col_rem_no,
                col_rem_amount=col_rem_amount,
                col_rem_lin_no=line_num,
                col_co_code=col_co_code,
                col_cust_no=cust_no, 
                col_rem_line_amount=fee_amount,
                col_rem_date=col_rem_date,
                )
            messages.success(request, 'Customer successfully receipted.')
            return redirect('Beezfees:remmittance_list')
        except Exception as e:
            messages.error(request, f'An error occurred while saving data: {str(e)}') 
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT DISTINCT Grade FROM invoice_list_view")
            grades = cursor.fetchall()

        if request.method == 'POST':
            grade_filter = request.POST.get('grade_filter')
            with connection.cursor() as cursor:
                if grade_filter:
                    cursor.execute("SELECT * FROM invoice_list_view WHERE Grade = %s", [grade_filter])
                else:
                    cursor.execute("SELECT * FROM invoice_list_view")
                invoice = cursor.fetchall()
        else:
            with connection.cursor() as cursor:
                cursor.execute("SELECT * FROM invoice_list_view")
                invoice = cursor.fetchall()
    
        return render(request, 'remit_exchange.html', {'invoice': invoice})
    except Exception as e:
        messages.error(request, f"An error occurred: {str(e)}")
        return render(request, 'remit_exchange.html', )
   # Fetch invoiced customers for rendering                   
    try:
        customers = lbt_invoice_hdr.objects.all()
        names = lbt_customer.objects.filter(col_cust_no__in=customers.values_list('col_cust_no', flat=True))
        currencies = lbt_currency.objects.filter(col_curr_code__in=customers.values_list('col_curr_code', flat=True))
        context = {
        'invoiced_list': customers,
        'names':  names,
        'currencies': currencies,
        'User':dd
    }
    
        return render(request, 'remit_exchange.html', context)
    except Exception as e:
            messages.error(request, f'An error occurred while retrieving data: {str(e)}') 
            
            

# from django.contrib.auth.hashers import make_password
            
# @login_required
# def add_users(request):
#     if request.method == 'POST':
#         users = None
        
#         try:
#             username = request.POST.get('username')
#             first_name = request.POST.get('first_name')
#             last_name = request.POST.get('last_name')
#             date_joined = request.POST.get('date_joined')
#             password = request.POST.get('password')
#             is_active = bool(request.POST.get('is_active'))
        
#             # Encrypt the password
#             encrypted_password = make_password(password)
            
#             users = User(
#                 username=username,
#                 first_name=first_name,
#                 last_name=last_name,
#                 date_joined=date_joined,
#                 is_active=is_active,
#                 password=encrypted_password
#             )
#             users.save()
            
#             messages.success(request, 'User successfully saved.')
#         except Exception as e:
#             messages.error(request, f'An error occurred while saving the user: {str(e)}')
        
#     try:
#         user = User.objects.all()
#     except Exception as e:
#         messages.error(request, f'An error occurred while retrieving users: {str(e)}')
#         user = []
    
#     context = {
#         'user': user,
#     }
    
#     return render(request, 'add_users.html', context)           



